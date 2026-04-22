import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import time
import warnings
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import json
import requests
import re
warnings.filterwarnings('ignore')

def formatear_numero(numero, decimales=2, usar_separador_miles=True):
    """
    Formatea un número con separadores de miles y decimales según formato argentino.
    Usa punto para miles y coma para decimales (ej: 1.234,56)
    """
    if numero is None or pd.isna(numero):
        return "-"
    
    try:
        numero = float(numero)
        
        if usar_separador_miles:
            # Formatear con separador de miles y decimales
            # Usar formato: 1.234,56
            parte_entera = int(abs(numero))
            parte_decimal = abs(numero) - parte_entera
            
            # Formatear parte entera con separador de miles
            parte_entera_str = f"{parte_entera:,}".replace(",", ".")
            
            # Formatear parte decimal
            if decimales > 0:
                parte_decimal_str = f"{parte_decimal:.{decimales}f}".split('.')[1]
                signo = "-" if numero < 0 else ""
                return f"{signo}{parte_entera_str},{parte_decimal_str}"
            else:
                signo = "-" if numero < 0 else ""
                return f"{signo}{parte_entera_str}"
        else:
            # Sin separador de miles, solo decimales
            if decimales > 0:
                return f"{numero:,.{decimales}f}".replace(",", "X").replace(".", ",").replace("X", ".")
            else:
                return f"{numero:,.0f}".replace(",", ".")
    except:
        return str(numero)


def _esc_html(v):
    """Escape special HTML characters in a string value."""
    return str(v).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;')


TABLA_BONOS_CSS = """
<style>
.bond-wrap { border-radius:10px; overflow:hidden; border:1px solid #e0e0e0; }
.bond-title { background:#fafafa; color:#333; font-weight:700; font-size:14px; padding:11px 14px; border-bottom:2px solid #e0e0e0; font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif; letter-spacing:0.02em; }
.bond-table { width:100%; border-collapse:collapse; font-size:13px; font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif; }
.bond-table th { background:#fafafa; color:#555; font-weight:600; padding:9px 12px; text-align:center; border-bottom:2px solid #e0e0e0; white-space:nowrap; }
.bond-table th:first-child { text-align:left; }
.bond-table td { padding:8px 12px; color:#333; white-space:nowrap; text-align:center; }
.bond-table td:first-child { text-align:left; }
.bond-table tr:nth-child(even) td { background:#f7f7f7; }
.bond-table tr:nth-child(odd) td { background:#ffffff; }
.bond-table tr:hover td { background:#eef2ff; }
</style>
"""


def render_tabla_bonos_html(df, titulo='', columnas_derecha=None, columnas_color_signo=None):
    columnas_derecha = columnas_derecha or []
    columnas_color_signo = columnas_color_signo or []

    cols = list(df.columns)
    headers = ''.join(
        f'<th style="text-align:{"right" if c in columnas_derecha else "left" if i == 0 else "center"}">{_esc_html(c)}</th>'
        for i, c in enumerate(cols)
    )

    rows = ''
    for _, row in df.iterrows():
        cells = ''
        for i, col in enumerate(cols):
            val = row[col]
            val_str = _esc_html(val)
            align = 'right' if col in columnas_derecha else 'left' if i == 0 else 'center'
            extra_style = ''
            if col in columnas_color_signo and val != '-':
                val_text = str(val).strip()
                if val_text.startswith('+'):
                    extra_style = 'color:#2e7d32;font-weight:600;'
                elif val_text.startswith('-'):
                    extra_style = 'color:#c62828;font-weight:600;'
            cells += f'<td style="text-align:{align};{extra_style}">{val_str}</td>'
        rows += f'<tr>{cells}</tr>'

    title_html = f'<div class="bond-title">{_esc_html(titulo)}</div>' if titulo else ''
    return f'<div class="bond-wrap">{title_html}<table class="bond-table"><thead><tr>{headers}</tr></thead><tbody>{rows}</tbody></table></div>'


# Configuración de la página
st.set_page_config(
    page_title="Calculadora de Bonos",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personalizado para el dashboard
st.markdown("""
<style>
    /* Forzar tema claro */
    .stApp {
        background-color: #f4f6fb !important;
        color: black !important;
    }

    .stApp > div {
        background-color: #f4f6fb !important;
    }
    
    /* Sidebar siempre visible - ocultar botón de colapsar */
    [data-testid="collapsedControl"] { display: none !important; }
    [data-testid="stSidebarCollapseButton"] { display: none !important; }
    [data-testid="stSidebar"] > div > button { display: none !important; }
    [data-testid="stSidebarNav"] { display: none !important; }
    button[kind="header"] { display: none !important; }
    .st-emotion-cache-1dp5vir, .st-emotion-cache-czk5ss { display: none !important; }
    [data-testid="stSidebar"] {
        display: flex !important;
        visibility: visible !important;
        transform: none !important;
        left: 0 !important;
        min-width: 280px !important;
        width: 280px !important;
        overflow-y: auto !important;
        overflow-x: hidden !important;
    }
    [data-testid="stSidebarContent"] {
        display: flex !important;
        width: 100% !important;
        overflow: visible !important;
        padding-top: 3.5rem !important;
    }
    /* Ocultar barra superior de Streamlit */
    [data-testid="stToolbar"] { display: none !important; }
    [data-testid="stDecoration"] { display: none !important; }
    [data-testid="stStatusWidget"] { display: none !important; }
    header[data-testid="stHeader"] {
        height: 0px !important;
        min-height: 0px !important;
        padding: 0 !important;
        overflow: hidden !important;
    }

    /* Selectores modernos Streamlit 1.x */
    [data-testid="stMainBlockContainer"] {
        padding-top: 0.75rem !important;
        padding-left: 2rem !important;
    }
    [data-testid="stAppViewContainer"] > section {
        padding-top: 0 !important;
    }

    /* Selectores legacy (versiones anteriores) */
    .main .block-container {
        padding-left: 2rem !important;
        padding-top: 0.75rem !important;
        padding-bottom: 1rem !important;
    }
    .main > div {
        padding-top: 0rem;
    }
    
    .main .element-container {
        margin-top: 0rem !important;
    }
    
    /* Eliminar espaciado del primer elemento */
    .main .element-container:first-child {
        margin-top: 0rem !important;
        padding-top: 0rem !important;
    }
    
    /* Selectores más específicos para eliminar espaciado */
    .main .block-container > div {
        padding-top: 0rem !important;
        margin-top: 0rem !important;
    }
    
    .main .block-container > div:first-child {
        padding-top: 0rem !important;
        margin-top: 0rem !important;
    }
    
    /* Eliminar espaciado de columnas */
    .main .stColumn {
        padding-top: 0rem !important;
        margin-top: 0rem !important;
    }
    
    .main .stColumn:first-child {
        padding-top: 0rem !important;
        margin-top: 0rem !important;
    }
    
    /* Eliminar espaciado de contenedores de Streamlit */
    div[data-testid="stMarkdownContainer"] {
        margin-top: 0rem !important;
        padding-top: 0rem !important;
    }
    
    div[data-testid="stMarkdownContainer"]:first-child {
        margin-top: 0rem !important;
        padding-top: 0rem !important;
    }
    
    /* Alinear sección de resultados con el título del sidebar */
    .main .stColumn:first-child {
        margin-top: -2rem !important;
    }
    
    /* Reducir espaciado entre elementos en la columna principal */
    .main .stColumn:first-child .element-container {
        margin-top: -1rem !important;
    }
    
    /* Ajustar espaciado de las tarjetas */
    .main .stColumn:first-child .metrics-grid {
        margin-top: -1.5rem !important;
    }
    
    /* Eliminar bordes de iframes de Streamlit */
    div[data-testid="stIFrame"] {
        border: none !important;
        box-shadow: none !important;
        outline: none !important;
    }
    
    /* Reducir espaciado del título "Datos de Mercado" */
    .main h2 {
        margin-bottom: 0.2rem !important;
        padding-bottom: 0.2rem !important;
        line-height: 1.2 !important;
    }
    
    /* Eliminar espaciado extra después de títulos h2 */
    .main h2 + .stMarkdown {
        margin-top: -0.5rem !important;
    }
    
    /* Reducir espaciado de elementos después de títulos */
    .main h2 + div {
        margin-top: -0.5rem !important;
    }
    
    
    
    /* Reducir tamaño de títulos y alinear arriba */
    .main h1, .main h2, .main h3, 
    .main .element-container h1, .main .element-container h2, .main .element-container h3,
    .main .block-container h1, .main .block-container h2, .main .block-container h3 {
        margin-top: 0rem !important;
        margin-bottom: 0.5rem !important;
        padding-top: 0rem !important;
        font-size: 1.2rem !important;
        line-height: 1.2 !important;
    }
    
    .main h1, .main .element-container h1, .main .block-container h1 {
        font-size: 1.4rem !important;
    }
    
    .main h2, .main .element-container h2, .main .block-container h2 {
        font-size: 1.2rem !important;
    }
    
    .main h3, .main .element-container h3, .main .block-container h3 {
        font-size: 1.1rem !important;
    }
    
    /* Selectores más específicos para Streamlit */
    div[data-testid="stMarkdownContainer"] h1,
    div[data-testid="stMarkdownContainer"] h2,
    div[data-testid="stMarkdownContainer"] h3 {
        margin-top: 0rem !important;
        margin-bottom: 0.5rem !important;
        padding-top: 0rem !important;
        font-size: 1.2rem !important;
        line-height: 1.2 !important;
    }
    
    
    .main .element-container {
        margin-top: 0.2rem !important;
        margin-bottom: 0.2rem !important;
    }
    
    /* Reducir espaciado entre elementos */
    .stSelectbox, .stDateInput, .stNumberInput, .stButton {
        margin-top: 0.1rem !important;
        margin-bottom: 0.1rem !important;
    }
    /* Reducir gap entre fecha y precio en la calculadora */
    [data-testid="stDateInput"] {
        margin-bottom: 0.3rem !important;
    }
    [data-testid="stNumberInput"] {
        margin-top: 0.3rem !important;
    }
    
    div[data-testid="stMarkdownContainer"] h1 {
        font-size: 1.4rem !important;
    }
    
    div[data-testid="stMarkdownContainer"] h2 {
        font-size: 1.2rem !important;
    }
    
    div[data-testid="stMarkdownContainer"] h3 {
        font-size: 1.1rem !important;
    }
    
    /* Sidebar */
    .stSidebar {
        background: linear-gradient(135deg, #94a3b8 0%, #64748b 100%);
        color: white;
    }
    
    /* Bajar el título del sidebar */
    .stSidebar h1 {
        margin-top: 0.8rem !important;
        padding-top: 0.8rem !important;
    }
    
    .stSidebar .element-container:first-child h1 {
        margin-top: 0.8rem !important;
        padding-top: 0.8rem !important;
    }
    
    .stSidebar .stSelectbox > div > div {
        background-color: rgba(255, 255, 255, 0.1);
        border: 1px solid rgba(255, 255, 255, 0.2);
        color: white;
    }
    
    .stSidebar .stSelectbox label {
        color: white !important;
        font-weight: 600;
    }
    
    .stSidebar .stDateInput > div > div {
        background-color: rgba(255, 255, 255, 0.1);
        border: 1px solid rgba(255, 255, 255, 0.2);
        color: white;
    }
    
    .stSidebar .stDateInput label {
        color: white !important;
        font-weight: 600;
    }
    
    .stSidebar .stDateInput > div > div > input {
        color: white !important;
    }
    
    .stSidebar .stDateInput > div > div > input::placeholder {
        color: rgba(255, 255, 255, 0.7) !important;
    }
    
    .stSidebar .stNumberInput > div > div > input {
        background-color: rgba(255, 255, 255, 0.1);
        border: 1px solid rgba(255, 255, 255, 0.2);
        color: white;
    }
    
    .stSidebar .stNumberInput label {
        color: white !important;
        font-weight: 600;
    }
    
    .stSidebar .stButton > button {
        background: linear-gradient(135deg, #3b82f6 0%, #1d4ed8 100%);
        color: white;
        border: none;
        border-radius: 8px;
        padding: 0.5rem 1rem;
        font-weight: 600;
        width: 100%;
    }
    
    .stSidebar .stButton > button:hover {
        background: linear-gradient(135deg, #1d4ed8 0%, #1e40af 100%);
        transform: translateY(-1px);
        box-shadow: 0 4px 12px rgba(59, 130, 246, 0.3);
    }
    
    /* Botones secondary (gris) */
    .stButton > button[kind="secondary"] {
        background-color: #6b7280 !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        padding: 0.5rem 1rem !important;
        font-weight: 600 !important;
    }

    .stButton > button[kind="secondary"]:hover {
        background-color: #4b5563 !important;
        transform: translateY(-1px) !important;
        box-shadow: 0 4px 12px rgba(107, 114, 128, 0.3) !important;
    }

    /* Botón primary — mismo azul que sidebar */
    .stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #3b82f6 0%, #1d4ed8 100%) !important;
        color: white !important;
        border: none !important;
        border-radius: 8px !important;
        padding: 0.5rem 1rem !important;
        font-weight: 600 !important;
    }

    .stButton > button[kind="primary"]:hover {
        background: linear-gradient(135deg, #1d4ed8 0%, #1e40af 100%) !important;
        transform: translateY(-1px) !important;
        box-shadow: 0 4px 12px rgba(59, 130, 246, 0.3) !important;
    }
    
    /* Ocultar botones +/- del number input */
    .stNumberInput button {
        display: none !important;
    }
    
    /* Columna izquierda: flex para que el card de info llene el espacio */
    [data-testid="stHorizontalBlock"] > [data-testid="stColumn"]:first-child > [data-testid="stVerticalBlock"] {
        display: flex !important;
        flex-direction: column !important;
        height: 100% !important;
    }


    /* Card contenedor de grupo de métricas */
    .metrics-card {
        background: white;
        border-radius: 14px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.07);
        padding: 1.2rem 1.2rem 1rem;
        margin-bottom: 2.2rem;
    }
    .metrics-card-title {
        font-size: 1.05rem;
        font-weight: 700;
        color: #94a3b8;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        margin-bottom: 0.9rem;
        padding-bottom: 0.6rem;
        border-bottom: 1px solid #f0f2f7;
    }

    /* Cards de métricas individuales */
    .metric-card {
        background: #f4f6fb;
        border-radius: 10px;
        padding: 0.85rem 0.6rem;
        text-align: center;
        display: flex !important;
        flex-direction: column !important;
        justify-content: center !important;
        align-items: center !important;
    }
    .metric-label {
        font-size: 1.05rem;
        font-weight: 600;
        color: #64748b;
        margin-bottom: 0.35rem;
        text-transform: none;
        letter-spacing: 0;
        text-align: center;
    }
    .metric-value {
        font-size: 1.05rem;
        font-weight: 700;
        color: #1a237e;
        line-height: 1.2;
        text-align: center;
    }

    /* Grid de métricas */
    .metrics-grid {
        display: grid;
        grid-template-columns: repeat(4, 1fr);
        gap: 0.6rem;
    }
    .metrics-row {
        display: grid;
        grid-template-columns: repeat(4, 1fr);
        gap: 0.6rem;
    }

    /* Card para inputs y para info del bono */
    .calc-card {
        background: white;
        border-radius: 14px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.07);
        padding: 1.2rem 1.2rem 1rem;
        margin-bottom: 1rem;
    }
    .calc-card-title {
        font-size: 0.93rem;
        font-weight: 700;
        color: #94a3b8;
        text-transform: uppercase;
        letter-spacing: 0.08em;
        margin-bottom: 0.9rem;
        padding-bottom: 0.6rem;
        border-bottom: 1px solid #f0f2f7;
    }
    /* Card info bono: ancho completo de la columna */
    .calc-card-fill {
        background: white;
        border-radius: 14px;
        box-shadow: 0 2px 10px rgba(0,0,0,0.07);
        padding: 0.9rem 1.2rem 0.7rem;
        flex: 1;
        width: 100%;
        box-sizing: border-box;
    }
    /* Forzar que el contenedor de Streamlit no agregue padding extra */
    [data-testid="stVerticalBlock"] > [data-testid="stVerticalBlockBorderWrapper"],
    [data-testid="stVerticalBlock"] > div > [data-testid="stMarkdownContainer"]:has(.calc-card-fill) {
        padding: 0 !important;
        margin: 0 !important;
        width: 100% !important;
    }
    
    /* Info bullets */
    .info-bullets {
        background: linear-gradient(135deg, #f1f5f9 0%, #e2e8f0 100%);
        border-radius: 8px;
        padding: 1rem;
        margin: 1rem 0;
    }
    
    .info-bullets h4 {
        color: #475569;
        margin-bottom: 0.5rem;
        font-size: 0.9rem;
    }
    
    .info-bullets p {
        color: #64748b;
        margin: 0.25rem 0;
        font-size: 0.85rem;
    }
    
    /* Future content */
    .future-content {
        background: linear-gradient(135deg, #f8fafc 0%, #e2e8f0 100%);
        border-radius: 12px;
        padding: 1.5rem;
        margin: 1rem 0;
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
    }
    
    .future-content h3 {
        color: #1e293b;
        margin-bottom: 1rem;
        font-size: 1.1rem;
    }
    
    .future-content ul {
        color: #64748b;
        margin: 0;
        padding-left: 1.5rem;
    }
    
    .future-content li {
        margin: 0.5rem 0;
        font-size: 0.9rem;
    }
    
    /* Cash flow table */
    .cashflow-table {
        background: white;
        border-radius: 8px;
        overflow: hidden;
        box-shadow: 0 2px 8px rgba(0, 0, 0, 0.1);
    }
    
    /* Ocultar bordes del dataframe */
    .stDataFrame {
        border: none !important;
    }
    
    .stDataFrame > div {
        border: none !important;
    }
    
    .stDataFrame table {
        border: none !important;
    }
    
    .stDataFrame th, .stDataFrame td {
        border: none !important;
        border-bottom: 1px solid #e2e8f0 !important;
    }
    
    .stDataFrame th {
        background-color: #f8fafc !important;
        font-weight: 600 !important;
        color: #475569 !important;
    }
    
    /* Ocultar menú del dataframe */
    .stDataFrame > div > div:first-child {
        display: none !important;
    }
    
    /* Ocultar barra de GitHub */
    .stDeployButton {
        display: none !important;
    }
    
    /* Ocultar elementos de GitHub/Streamlit Cloud */
    [data-testid="stDeployButton"] {
        display: none !important;
    }
    
    /* Ocultar cualquier elemento de deploy */
    .stDeployButton,
    .deploy-button,
    .github-button {
        display: none !important;
    }
    
    /* Ocultar barra superior de Streamlit Cloud */
    .stApp > header {
        display: none !important;
    }
    
    /* Ocultar elementos de la barra superior */
    a[href*="github"],
    a[href*="deploy"],
    button[title*="Deploy"],
    button[title*="GitHub"] {
        display: none !important;
    }
    
    /* Reducir tamaño de fuentes en widgets de TradingView */
    iframe[src*="tradingview"] {
        font-size: 10px !important;
    }
    
    /* Reducir tamaño de fuentes en contenedores de TradingView */
    div[data-testid="stIFrame"] iframe {
        font-size: 10px !important;
    }
    
    /* Reducir tamaño de fuentes en elementos de TradingView */
    .tradingview-widget-container {
        font-size: 10px !important;
    }
    
    .tradingview-widget-container * {
        font-size: 10px !important;
    }
    
    /* Reducir tamaño de fuentes en iframes de TradingView */
    iframe[title*="TradingView"] {
        font-size: 10px !important;
    }
    
    /* Reducir tamaño de fuentes en todos los iframes */
    iframe {
        font-size: 10px !important;
    }
    
    /* Reducir tamaño de fuentes específicamente en el widget Market Data */
    .tradingview-widget-container[style*="height: 800px"] {
        font-size: 8px !important;
    }
    
    .tradingview-widget-container[style*="height: 800px"] * {
        font-size: 8px !important;
    }
    
    /* Reducir tamaño de fuentes en el widget Market Data */
    div[data-testid="stIFrame"] iframe[src*="market-overview"] {
        font-size: 8px !important;
    }

    /* Labels de inputs visibles */
    [data-testid="stDateInput"] label,
    [data-testid="stNumberInput"] label,
    [data-testid="stWidgetLabel"],
    .stDateInput label,
    .stNumberInput label {
        color: #374151 !important;
        font-size: 0.85rem !important;
        font-weight: 600 !important;
    }
    /* Inputs: fondo blanco, borde suave */
    [data-testid="stDateInput"] input,
    [data-testid="stNumberInput"] input {
        background-color: white !important;
        color: #1a1a1a !important;
        border: 1px solid #dee2e6 !important;
        border-radius: 8px !important;
    }
    [data-baseweb="input"], [data-baseweb="base-input"] {
        border-radius: 8px !important;
    }
    /* Ocultar "Press Enter to apply" en number inputs */
    [data-testid="stNumberInput"] ~ small,
    [data-testid="stNumberInput"] small,
    [data-testid="stNumberInput"] + div small {
        display: none !important;
    }
    /* Color azul en las tabs */
    [data-testid="stTabs"] button[role="tab"] {
        color: #1a56db !important;
        background-color: #e4e6ea !important;
    }
    [data-testid="stTabs"] [data-baseweb="tab-list"] button[role="tab"] {
        padding: 8px 20px !important;
        border-radius: 6px 6px 0 0 !important;
        font-size: 14px !important;
        font-weight: 700 !important;
        margin: 0 !important;
        border: 1px solid #b8bcc4 !important;
        border-bottom: none !important;
    }
    [data-testid="stTabs"] [data-baseweb="tab-list"] {
        gap: 0 !important;
        border-bottom: 2px solid #e4e6ea !important;
    }
    [data-testid="stTabs"] [data-baseweb="tab-list"] button[role="tab"] p,
    [data-testid="stTabs"] [data-baseweb="tab-list"] button[role="tab"] div,
    [data-testid="stTabs"] [data-baseweb="tab-list"] button[role="tab"] span {
        font-weight: 700 !important;
        font-size: 14px !important;
    }
    [data-testid="stTabs"] [data-baseweb="tab-list"] button[role="tab"][aria-selected="true"] {
        color: #1a56db !important;
        background-color: #d1d5db !important;
        border-radius: 6px 6px 0 0 !important;
    }
    /* Asegurar que el panel de contenido no herede el gris */
    [data-testid="stTabs"] [data-baseweb="tab-panel"] {
        background-color: transparent !important;
    }
    /* Línea de selección azul (solo el indicador activo, no la línea base) */
    [data-testid="stTabs"] [data-baseweb="tab-highlight"] {
        background-color: #1a56db !important;
    }
</style>
""", unsafe_allow_html=True)

# JS para reducir padding-top del bloque principal (emotion CSS-in-JS se inyecta tarde)
st.markdown("""
<script>
(function fixStyles() {
    function applyFix() {
        // Padding superior
        var el = document.querySelector('[data-testid="stMainBlockContainer"]');
        if (el) el.style.setProperty('padding-top', '0.25rem', 'important');
        var header = document.querySelector('header[data-testid="stHeader"]');
        if (header) {
            header.style.setProperty('height', '0px', 'important');
            header.style.setProperty('min-height', '0px', 'important');
            header.style.setProperty('overflow', 'hidden', 'important');
        }
        // Inputs: fondo blanco, texto oscuro
        document.querySelectorAll('[data-baseweb="input"], [data-baseweb="base-input"]').forEach(function(el) {
            el.style.setProperty('background-color', 'white', 'important');
        });
        document.querySelectorAll('[data-testid="stDateInput"] input, [data-testid="stNumberInput"] input').forEach(function(el) {
            el.style.setProperty('background-color', 'white', 'important');
            el.style.setProperty('color', '#1a1a1a', 'important');
        });
        // Labels visibles
        document.querySelectorAll('[data-testid="stDateInput"] label, [data-testid="stNumberInput"] label, [data-testid="stWidgetLabel"]').forEach(function(el) {
            el.style.setProperty('color', '#374151', 'important');
            el.style.setProperty('font-weight', '600', 'important');
        });
        // Forzar ancho completo en toda la cadena sobre .calc-card-fill
        document.querySelectorAll('.calc-card-fill').forEach(function(el) {
            el.style.setProperty('width', '100%', 'important');
            el.style.setProperty('box-sizing', 'border-box', 'important');
            var node = el.parentElement;
            var steps = 0;
            while (node && steps < 8) {
                if (node.matches('[data-testid="stColumn"]')) break;
                node.style.setProperty('padding-left', '0', 'important');
                node.style.setProperty('padding-right', '0', 'important');
                node.style.setProperty('width', '100%', 'important');
                node.style.setProperty('max-width', '100%', 'important');
                node = node.parentElement;
                steps++;
            }
        });
    }
    if (document.readyState === 'loading') {
        document.addEventListener('DOMContentLoaded', function() { setTimeout(applyFix, 100); });
    } else {
        setTimeout(applyFix, 100);
    }
    setTimeout(applyFix, 500);
    setTimeout(applyFix, 1500);


    // Reaplica cada vez que Streamlit hace rerender
    var _fixTimer = null;
    var observer = new MutationObserver(function() {
        if (_fixTimer) clearTimeout(_fixTimer);
        _fixTimer = setTimeout(applyFix, 150);
    });
    observer.observe(document.body, { childList: true, subtree: true });
})();
</script>
""", unsafe_allow_html=True)


# ── Feriados Argentina 2025-2026 ─────────────────────────────────────────────
import datetime as _dt
FERIADOS_ARGENTINA = {
    # 2025
    _dt.date(2025,  1,  1), _dt.date(2025,  3,  3), _dt.date(2025,  3,  4),
    _dt.date(2025,  3, 24), _dt.date(2025,  4,  2), _dt.date(2025,  4, 18),
    _dt.date(2025,  5,  1), _dt.date(2025,  5, 25), _dt.date(2025,  6, 20),
    _dt.date(2025,  7,  9), _dt.date(2025,  8, 18), _dt.date(2025, 10, 13),
    _dt.date(2025, 11, 24), _dt.date(2025, 12,  8), _dt.date(2025, 12, 25),
    # 2026
    _dt.date(2026,  1,  1), _dt.date(2026,  2, 16), _dt.date(2026,  2, 17),
    _dt.date(2026,  3, 23), _dt.date(2026,  3, 24), _dt.date(2026,  4,  2), _dt.date(2026,  4,  3),
    _dt.date(2026,  5,  1), _dt.date(2026,  5, 25), _dt.date(2026,  6, 20),
    _dt.date(2026,  7,  9), _dt.date(2026,  8, 17), _dt.date(2026, 10, 12),
    _dt.date(2026, 11, 23), _dt.date(2026, 12,  8), _dt.date(2026, 12, 25),
}

def es_dia_habil(fecha):
    """Devuelve True si la fecha es día hábil (no fin de semana ni feriado)."""
    if hasattr(fecha, 'date'):
        fecha = fecha.date()
    return fecha.weekday() < 5 and fecha not in FERIADOS_ARGENTINA

def get_next_business_day():
    """Devuelve el próximo día hábil desde hoy, considerando feriados."""
    d = datetime.now().date() + timedelta(days=1)
    while not es_dia_habil(d):
        d += timedelta(days=1)
    return datetime(d.year, d.month, d.day)

def n_dias_habiles_antes(fecha, n):
    """Devuelve la fecha que cae n días hábiles antes de la fecha dada."""
    if hasattr(fecha, 'date'):
        fecha = fecha.date()
    d = fecha - timedelta(days=1)
    count = 0
    while count < n:
        if es_dia_habil(d):
            count += 1
        if count < n:
            d -= timedelta(days=1)
    return d

# ── CER desde BCRA ────────────────────────────────────────────────────────────
def _cer_periodo_actual():
    """
    Clave de período que cambia solo el día 15 de cada mes,
    que es cuando el BCRA publica el nuevo XLS con los valores del mes.
    Antes del 15: clave = mes anterior. A partir del 15: clave = mes actual.
    """
    hoy = datetime.now().date()
    if hoy.day >= 15:
        return (hoy.year, hoy.month)
    else:
        primer_dia_mes = hoy.replace(day=1)
        mes_anterior = primer_dia_mes - timedelta(days=1)
        return (mes_anterior.year, mes_anterior.month)

@st.cache_data
def obtener_cer_bcra(periodo=None):
    """
    Descarga los valores diarios de CER desde los XLS anuales del BCRA.
    URL: https://www.bcra.gob.ar/pdfs/PublicacionesEstadisticas/cerYYYY.xls
    El parámetro `periodo` actúa como cache key: solo cambia el día 15 de cada
    mes, por lo que el XLS se descarga una sola vez por período mensual.
    Retorna dict {datetime.date: float}.
    """
    try:
        import xlrd, ssl
        ctx = ssl._create_unverified_context()
    except ImportError:
        return {}

    cer_dict = {}
    current_year = datetime.now().year
    for year in [current_year - 1, current_year]:
        url = f"https://www.bcra.gob.ar/pdfs/PublicacionesEstadisticas/cer{year}.xls"
        try:
            import urllib.request
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=15, context=ctx) as r:
                data = r.read()
            wb = xlrd.open_workbook(file_contents=data)
            sh = wb.sheets()[0]
            for i in range(sh.nrows):
                if sh.cell_type(i, 4) == 2 and sh.cell_type(i, 5) == 2:
                    fecha_str = str(int(sh.cell_value(i, 4)))
                    if len(fecha_str) == 8:
                        fecha = _dt.date(int(fecha_str[:4]), int(fecha_str[4:6]), int(fecha_str[6:8]))
                        cer_dict[fecha] = sh.cell_value(i, 5)
        except Exception:
            pass
    return cer_dict

def obtener_cer_settlement(fecha_liquidacion):
    """
    Retorna (valor_cer, fecha_cer) donde fecha_cer es el 10° día hábil
    anterior a fecha_liquidacion, según la convención de bonos CER.
    """
    cer_data = obtener_cer_bcra(periodo=_cer_periodo_actual())
    if not cer_data:
        return None, None
    if hasattr(fecha_liquidacion, 'date'):
        fecha_liquidacion = fecha_liquidacion.date()
    fecha_cer = n_dias_habiles_antes(fecha_liquidacion, 10)
    return cer_data.get(fecha_cer), fecha_cer

# Función para calcular días entre fechas según base de cálculo
def calcular_dias(fecha1, fecha2, base_calculo):
    # Convertir ambas fechas a datetime para asegurar compatibilidad
    if hasattr(fecha1, 'date'):
        fecha1 = fecha1.date()
    if hasattr(fecha2, 'date'):
        fecha2 = fecha2.date()
    
    if base_calculo == "30/360":
        d1, m1, y1 = fecha1.day, fecha1.month, fecha1.year
        d2, m2, y2 = fecha2.day, fecha2.month, fecha2.year
        
        # Ajuste para 30/360
        if d1 == 31:
            d1 = 30
        if d2 == 31 and d1 == 30:
            d2 = 30
        
        dias = (y2 - y1) * 360 + (m2 - m1) * 30 + (d2 - d1)
        return dias
    elif base_calculo == "ACT/360":
        return (fecha2 - fecha1).days
    elif base_calculo == "ACT/365":
        return (fecha2 - fecha1).days
    elif base_calculo == "ACT/ACT":
        return (fecha2 - fecha1).days
    else:
        return (fecha2 - fecha1).days

# Función para calcular YTM usando Newton-Raphson
def calcular_ytm(precio_dirty, flujos, fechas, fecha_liquidacion, base_calculo="ACT/365", periodicidad=2):
    # Convertir fecha_liquidacion a date si es datetime
    fecha_liq = fecha_liquidacion
    if hasattr(fecha_liq, 'date'):
        fecha_liq = fecha_liq.date()
    
    def npv(rate):
        total = 0
        for i, (flujo, fecha) in enumerate(zip(flujos, fechas)):
            dias = calcular_dias(fecha_liq, fecha, base_calculo)
            if base_calculo == "30/360":
                factor_descuento = (1 + rate) ** (dias / 360)
            elif base_calculo == "ACT/360":
                factor_descuento = (1 + rate) ** (dias / 360)
            elif base_calculo == "ACT/365":
                factor_descuento = (1 + rate) ** (dias / 365)
            elif base_calculo == "ACT/ACT":
                factor_descuento = (1 + rate) ** (dias / 365)
            else:
                factor_descuento = (1 + rate) ** (dias / 365)
            total += flujo / factor_descuento
        return total - precio_dirty
    
    def npv_derivative(rate):
        total = 0
        for i, (flujo, fecha) in enumerate(zip(flujos, fechas)):
            dias = calcular_dias(fecha_liq, fecha, base_calculo)
            if base_calculo == "30/360":
                factor_descuento = (1 + rate) ** (dias / 360)
                derivada = -flujo * (dias / 360) * (1 + rate) ** (dias / 360 - 1)
            elif base_calculo == "ACT/360":
                factor_descuento = (1 + rate) ** (dias / 360)
                derivada = -flujo * (dias / 360) * (1 + rate) ** (dias / 360 - 1)
            elif base_calculo == "ACT/365":
                factor_descuento = (1 + rate) ** (dias / 365)
                derivada = -flujo * (dias / 365) * (1 + rate) ** (dias / 365 - 1)
            elif base_calculo == "ACT/ACT":
                factor_descuento = (1 + rate) ** (dias / 365)
                derivada = -flujo * (dias / 365) * (1 + rate) ** (dias / 365 - 1)
            else:
                factor_descuento = (1 + rate) ** (dias / 365)
                derivada = -flujo * (dias / 365) * (1 + rate) ** (dias / 365 - 1)
            total += derivada
        return total
    
    # Método de Newton-Raphson
    rate = 0.05  # Tasa inicial
    tolerance = 1e-8
    max_iterations = 100
    
    for i in range(max_iterations):
        npv_val = npv(rate)
        if abs(npv_val) < tolerance:
            break
    
        derivative = npv_derivative(rate)
        if abs(derivative) < 1e-12:
            break
    
        rate = rate - npv_val / derivative
    
    # Si no converge, usar búsqueda binaria
    if abs(npv(rate)) > tolerance:
        low, high = 0.0, 1.0
        for _ in range(100):
            mid = (low + high) / 2
            if npv(mid) > 0:
                low = mid
            else:
                high = mid
            if high - low < tolerance:
                break
        rate = (low + high) / 2
    
    return rate

# Función para calcular duración Macaulay
def calcular_duracion_macaulay(flujos, fechas, fecha_liquidacion, ytm, base_calculo="ACT/365"):
    # Convertir fecha_liquidacion a date si es datetime
    fecha_liq = fecha_liquidacion
    if hasattr(fecha_liq, 'date'):
        fecha_liq = fecha_liq.date()
    
    pv_total = 0
    pv_weighted = 0
    
    for flujo, fecha in zip(flujos, fechas):
        dias = calcular_dias(fecha_liq, fecha, base_calculo)
        if base_calculo == "30/360":
            factor_descuento = (1 + ytm) ** (dias / 360)
        elif base_calculo == "ACT/360":
            factor_descuento = (1 + ytm) ** (dias / 360)
        elif base_calculo == "ACT/365":
            factor_descuento = (1 + ytm) ** (dias / 365)
        elif base_calculo == "ACT/ACT":
            factor_descuento = (1 + ytm) ** (dias / 365)
        else:
            factor_descuento = (1 + ytm) ** (dias / 365)
        
        divisor_tiempo = 360 if base_calculo in ("30/360", "ACT/360") else 365
        pv = flujo / factor_descuento
        pv_total += pv
        pv_weighted += pv * (dias / divisor_tiempo)

    if pv_total == 0:
        return 0

    return pv_weighted / pv_total

# Función para calcular duración modificada
def calcular_duracion_modificada(duracion_macaulay, ytm, periodicidad):
    return duracion_macaulay / (1 + ytm / periodicidad)

# Función para calcular intereses corridos
def calcular_intereses_corridos(fecha_liquidacion, fecha_ultimo_cupon, tasa_cupon, capital_residual, base_calculo="ACT/365"):
    # Convertir fecha_liquidacion a date si es datetime
    fecha_liq = fecha_liquidacion
    if hasattr(fecha_liq, 'date'):
        fecha_liq = fecha_liq.date()
    
    dias = calcular_dias(fecha_ultimo_cupon, fecha_liq, base_calculo)
    if base_calculo == "30/360":
        return (tasa_cupon * capital_residual) / 360 * dias
    elif base_calculo == "ACT/360":
        return (tasa_cupon * capital_residual) / 360 * dias
    elif base_calculo == "ACT/365":
        return (tasa_cupon * capital_residual) / 365 * dias
    elif base_calculo == "ACT/ACT":
        return (tasa_cupon * capital_residual) / 365 * dias
    else:
        return (tasa_cupon * capital_residual) / 365 * dias

# Función para encontrar el último cupón
def encontrar_ultimo_cupon(fecha_liquidacion, fechas_cupones, fecha_emision=None):
    fechas_anteriores = [fecha for fecha in fechas_cupones if fecha <= fecha_liquidacion]
    if not fechas_anteriores:
        # Bono en primer período: usar fecha de emisión como referencia
        return fecha_emision
    return max(fechas_anteriores)

# Función para calcular vida media
def calcular_vida_media(flujos_capital, fechas, fecha_liquidacion, base_calculo="ACT/365"):
    # Convertir fecha_liquidacion a date si es datetime
    fecha_liq = fecha_liquidacion
    if hasattr(fecha_liq, 'date'):
        fecha_liq = fecha_liq.date()
    
    if not flujos_capital or sum(flujos_capital) == 0:
        return 0
    
    total_capital = sum(flujos_capital)
    vida_media = 0
    
    for flujo, fecha in zip(flujos_capital, fechas):
        if flujo > 0:
            dias = calcular_dias(fecha_liq, fecha, base_calculo)
            peso = flujo / total_capital
            vida_media += peso * (dias / 365)
    
    return vida_media

# Función para encontrar el próximo cupón
def encontrar_proximo_cupon(fecha_liquidacion, fechas_cupones):
    fechas_futuras = [fecha for fecha in fechas_cupones if fecha > fecha_liquidacion]
    if not fechas_futuras:
        return None
    return min(fechas_futuras)

# Función para encontrar el cupón vigente
def encontrar_cupon_vigente(fecha_liquidacion, flujos):
    """
    Encuentra el cupón vigente basado en la fecha de liquidación.
    Busca la fila cuya fecha es la inmediatamente anterior a la fecha de liquidación.
    """
    # Convertir fecha_liquidacion a date si es datetime
    fecha_liq = fecha_liquidacion
    if hasattr(fecha_liq, 'date'):
        fecha_liq = fecha_liq.date()
    
    # Filtrar fechas anteriores o iguales a la fecha de liquidación
    fechas_anteriores = []
    for flujo in flujos:
        fecha_flujo = flujo['fecha']
        if hasattr(fecha_flujo, 'date'):
            fecha_flujo = fecha_flujo.date()
        
        if fecha_flujo <= fecha_liq:
            fechas_anteriores.append((fecha_flujo, flujo['cupon_vigente']))
    
    if not fechas_anteriores:
        # Bono en su primer período: devolver la tasa del próximo cupón
        fechas_futuras = []
        for flujo in flujos:
            fecha_flujo = flujo['fecha']
            if hasattr(fecha_flujo, 'date'):
                fecha_flujo = fecha_flujo.date()
            if fecha_flujo > fecha_liq:
                fechas_futuras.append((fecha_flujo, flujo['cupon_vigente']))
        if fechas_futuras:
            return min(fechas_futuras, key=lambda x: x[0])[1]
        return 0.0

    # Encontrar la fecha más cercana (inmediatamente anterior)
    fecha_mas_cercana = max(fechas_anteriores, key=lambda x: x[0])
    return fecha_mas_cercana[1]

# Función para encontrar la fecha de vencimiento
def encontrar_fecha_vencimiento(flujos):
    """
    Encuentra la fecha de vencimiento del bono (última fecha de los flujos).
    """
    if not flujos:
        return None
    
    # Encontrar la fecha máxima entre todos los flujos
    fechas = [flujo['fecha'] for flujo in flujos]
    fecha_vencimiento = max(fechas)
    
    return fecha_vencimiento

@st.cache_data(ttl=60)  # Cache por 1 minuto
def obtener_precios_data912(endpoint):
    """
    Obtiene todos los precios desde data912.com para un endpoint dado.
    endpoint: 'arg_bonds' para soberanos USD, 'arg_corp' para corporativos
    Retorna dict {symbol: last_price} o {} si falla
    """
    try:
        url = f"https://data912.com/live/{endpoint}"
        response = requests.get(url, timeout=5)
        if response.status_code == 200:
            data = response.json()
            return {
                item['symbol']: {'c': item['c'], 'pct_change': item.get('pct_change')}
                for item in data
                if item.get('symbol') and item.get('c') and item['c'] > 0
            }
    except Exception:
        pass
    return {}

def obtener_precio_data912(ticker):
    """
    Obtiene el último precio operado de un bono desde data912.com.
    Busca en soberanos (arg_bonds) y corporativos (arg_corp).
    Retorna el precio (float) o None si no se encuentra.
    """
    if not ticker or ticker.strip() == '':
        return None

    ticker_clean = ticker.strip().upper()

    # Si el ticker tiene exactamente 4 caracteres, agregar 'D' al final
    if len(ticker_clean) == 4:
        ticker_clean = ticker_clean + 'D'

    for endpoint in ['arg_bonds', 'arg_corp']:
        precios = obtener_precios_data912(endpoint)
        if ticker_clean in precios:
            return round(float(precios[ticker_clean]['c']), 2)

    return None


def obtener_tipo_cambio_implicito_data912(tipo_cambio='Tipo de Cambio MEP'):
    """
    Calcula el tipo de cambio implícito usando AL30 contra su especie en USD/CCL.
    - MEP: AL30 / AL30D
    - CCL: AL30 / AL30C
    Retorna el valor (float) o None si no se puede calcular.
    """
    try:
        precios = obtener_precios_data912('arg_bonds')
        precio_al30 = precios.get('AL30', {}).get('c')
        ticker_ref = 'AL30D' if tipo_cambio == 'Tipo de Cambio MEP' else 'AL30C'
        precio_ref = precios.get(ticker_ref, {}).get('c')

        if precio_al30 and precio_ref and precio_ref > 0:
            return round(float(precio_al30) / float(precio_ref), 2)
    except Exception:
        pass

    return None


def _slugify_monitor(valor):
    return ''.join(ch.lower() if ch.isalnum() else '_' for ch in str(valor)).strip('_')


def _tabla_manual_id(tipo_bono):
    return f"tabla_{_slugify_monitor(tipo_bono)}"


def _manual_price_state_key(tabla_id, row_id):
    return f"{tabla_id}__precio_manual__{_slugify_monitor(row_id)}"


def normalizar_precio_manual_monitor(valor):
    if valor is None:
        return None
    try:
        if pd.isna(valor):
            return None
    except TypeError:
        pass
    if isinstance(valor, str):
        valor = valor.strip().replace(',', '.')
        if not valor:
            return None
    try:
        precio = float(valor)
    except (TypeError, ValueError):
        return None
    return round(precio, 6) if precio > 0 else None


def obtener_precio_manual_monitor(tipo_bono, row_id):
    tabla_id = _tabla_manual_id(tipo_bono)
    return normalizar_precio_manual_monitor(st.session_state.get(_manual_price_state_key(tabla_id, row_id)))


# Cargar datos del Excel
try:
    import openpyxl as _openpyxl

    # Época de Excel para convertir seriales numéricos (Corp ARG usa enteros en lugar de datetime)
    _EXCEL_EPOCH = datetime(1899, 12, 30)

    def _serial_to_dt(serial):
        """Convierte serial numérico de Excel a datetime."""
        return _EXCEL_EPOCH + timedelta(days=int(serial))

    def _sf(v, default=0.0):
        """Convierte a float de forma segura."""
        if v is None:
            return default
        try:
            f = float(v)
            return default if (f != f) else f  # NaN check
        except (ValueError, TypeError):
            return default

    def _infer_periodicidad(flujos):
        """Infiere la periodicidad a partir del espaciado entre fechas de flujos."""
        fechas = [f['fecha'] for f in flujos]
        if len(fechas) < 2:
            return 2
        diffs = []
        for i in range(min(6, len(fechas) - 1)):
            d1 = fechas[i].date() if hasattr(fechas[i], 'date') else fechas[i]
            d2 = fechas[i+1].date() if hasattr(fechas[i+1], 'date') else fechas[i+1]
            diffs.append((d2 - d1).days)
        avg = sum(diffs) / len(diffs)
        if avg < 45:  return 12  # mensual
        if avg < 100: return 4   # trimestral
        if avg < 200: return 2   # semestral
        return 1                 # anual

    # Configuración por hoja: tipo de bono y base de cálculo por defecto
    SHEET_CONFIG = {
        'Corp ARG': {'tipo_bono': 'Corporativo Ley ARG', 'base_default': 'ACT/365'},
        'Corp NY':  {'tipo_bono': 'Corporativo Ley NY',  'base_default': '30/360'},
        'Sov USD':  {'tipo_bono': 'Soberano USD',        'base_default': '30/360'},
    }

    wb = _openpyxl.load_workbook('flujos new.xlsx', data_only=True)
    bonos = []

    for sheet_name, config in SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        tipo_bono    = config['tipo_bono']
        base_default = config['base_default']
        current_bono = None
        skip_emission_row = False

        for row in ws.iter_rows(values_only=True):
            v0 = row[0]

            # --- Fila de nombre de bono: col0 es string no vacío ---
            if isinstance(v0, str) and v0.strip() and not v0.strip().startswith('#'):
                if current_bono and current_bono['flujos']:
                    bonos.append(current_bono)

                nombre = v0.strip()
                v1 = row[1]  # base_calculo override (si contiene '/')
                v3 = row[3]  # ticker (Corp) o tasa_cupon inicial (Sov USD)

                # base_calculo: override explícito si col1 contiene '/', si no default de hoja
                if isinstance(v1, str) and '/' in v1:
                    base_calculo = v1.strip()
                else:
                    base_calculo = base_default

                # ticker: para Sov USD el nombre ES el ticker; para Corp viene en col3
                if sheet_name == 'Sov USD':
                    ticker = nombre
                elif isinstance(v3, str) and v3.strip() and v3.strip() != '#N/A':
                    ticker = v3.strip()
                else:
                    ticker = ''

                current_bono = {
                    'nombre': nombre,
                    'base_calculo': base_calculo,
                    'tipo_bono': tipo_bono,
                    'tasa_cupon': 0.0,
                    'ticker': ticker,
                    'periodicidad': 2,  # se recalcula tras parsear los flujos
                    'fecha_emision': None,  # se llena con la fila siguiente
                    'flujos': []
                }
                skip_emission_row = True  # la siguiente fila es la fecha de emisión
                continue

            # --- Fila de fecha de emisión: guardar y saltar ---
            if skip_emission_row:
                skip_emission_row = False
                if current_bono and isinstance(v0, datetime):
                    current_bono['fecha_emision'] = v0
                continue

            if current_bono is None:
                continue

            # --- Fila de flujo ---
            # col0 puede ser datetime (Corp NY / Sov USD) o entero serial (Corp ARG)
            if isinstance(v0, datetime):
                fecha = v0
            elif isinstance(v0, (int, float)) and v0 > 1000:
                fecha = _serial_to_dt(v0)
            else:
                continue  # fila vacía o inválida

            tasa      = _sf(row[1])
            intereses = _sf(row[2])
            capital   = _sf(row[3])
            total     = intereses + capital

            if len(current_bono['flujos']) == 0:
                current_bono['tasa_cupon'] = tasa

            current_bono['flujos'].append({
                'fecha': fecha,
                'cupon': intereses,
                'capital': capital,
                'total': total,
                'cupon_vigente': tasa,
            })

        # No olvidar el último bono de la hoja
        if current_bono and current_bono['flujos']:
            bonos.append(current_bono)

    # Inferir periodicidad para cada bono (sólo los que tienen flujos)
    for bono in bonos:
        bono['periodicidad'] = _infer_periodicidad(bono['flujos'])

    # --- Parsear hoja Lecap (estructura diferente: bloques de 23 filas) ---
    if 'Lecap' in wb.sheetnames:
        ws_lecap = wb['Lecap']
        lrows = list(ws_lecap.iter_rows(values_only=True))
        LECAP_BLOCK = 23  # filas por título
        LECAP_START = 2   # índice 0-based de la primera fila de datos (fila 3)
        idx = LECAP_START
        while idx < len(lrows):
            row = lrows[idx]
            nombre = row[0]
            # Detectar fila de nombre: string que no es separador ni etiqueta
            if not isinstance(nombre, str) or nombre in ('--', 'LECAP') or idx + LECAP_BLOCK > len(lrows):
                idx += 1
                continue
            maturity   = row[1] if isinstance(row[1], datetime) else None
            tasa_cupon = _sf(row[2])
            # Issue Date está en row+2, col B
            issue_row  = lrows[idx + 2] if idx + 2 < len(lrows) else (None,) * 4
            fecha_emision = issue_row[1] if isinstance(issue_row[1], datetime) else None
            # Freq. Coupon en row+5, col B
            freq_row    = lrows[idx + 5] if idx + 5 < len(lrows) else (None,) * 4
            periodicidad = int(_sf(freq_row[1], 1)) if freq_row[1] is not None else 1
            # Count Days en row+6, col B
            count_row   = lrows[idx + 6] if idx + 6 < len(lrows) else (None,) * 4
            base_calculo = count_row[1] if isinstance(count_row[1], str) and '/' in count_row[1] else '30/360'
            # Sink Factor (Valor Final) en row+8, col B
            sink_row       = lrows[idx + 8] if idx + 8 < len(lrows) else (None,) * 4
            valor_final    = _sf(sink_row[1])
            # Días Remanente en row+11, col B
            remanente_row  = lrows[idx + 11] if idx + 11 < len(lrows) else (None,) * 4
            dias_remanente = _sf(remanente_row[1])
            # Vida Media en row+16, col E (Avg. Life)
            vida_row       = lrows[idx + 16] if idx + 16 < len(lrows) else (None,) * 22
            vida_media_lec = _sf(vida_row[4])
            # Duración Modificada en row+20, col B
            dur_row        = lrows[idx + 20] if idx + 20 < len(lrows) else (None,) * 4
            duracion_mod_lec = _sf(dur_row[1])

            # Filtrar tickers que empiezan con TT o TY
            if nombre.strip().upper().startswith(('TT', 'TY')):
                idx += LECAP_BLOCK
                continue

            bonos.append({
                'nombre':        nombre.strip(),
                'tipo_bono':     'Lecaps & Boncaps',
                'ticker':        nombre.strip(),
                'tasa_cupon':    tasa_cupon,
                'base_calculo':  base_calculo,
                'periodicidad':  periodicidad,
                'fecha_emision': fecha_emision,
                'maturity':      maturity,
                'valor_final':   valor_final,
                'dias_remanente': dias_remanente,
                'vida_media_lec': vida_media_lec,
                'duracion_mod_lec': duracion_mod_lec,
                'flujos':        [],   # cálculos pendientes de implementar
            })
            idx += LECAP_BLOCK

    # --- Parsear hoja CER (misma estructura de bloques de 23 filas que Lecap) ---
    if 'CER' in wb.sheetnames:
        ws_cer = wb['CER']
        crows = list(ws_cer.iter_rows(values_only=True))
        CER_BLOCK = 23
        CER_START = 2
        idx = CER_START
        while idx < len(crows):
            row = crows[idx]
            nombre = row[0]
            if not isinstance(nombre, str) or nombre.strip() in ('--', 'CER') or idx + CER_BLOCK > len(crows):
                idx += 1
                continue
            nombre = nombre.strip()
            maturity   = row[1] if isinstance(row[1], datetime) else None
            tasa_cupon = _sf(row[2])
            issue_row  = crows[idx + 2] if idx + 2 < len(crows) else (None,)*4
            fecha_emision = issue_row[1] if isinstance(issue_row[1], datetime) else None
            freq_row   = crows[idx + 5] if idx + 5 < len(crows) else (None,)*4
            periodicidad = int(_sf(freq_row[1], 2)) if freq_row[1] is not None else 2
            count_row  = crows[idx + 6] if idx + 6 < len(crows) else (None,)*4
            base_calculo = count_row[1] if isinstance(count_row[1], str) and '/' in count_row[1] else '30/360'
            resid_row  = crows[idx + 7] if idx + 7 < len(crows) else (None,)*4
            residual   = _sf(resid_row[1])
            sink_row   = crows[idx + 8] if idx + 8 < len(crows) else (None,)*4
            sink_factor = _sf(sink_row[1])
            fcer_row   = crows[idx + 9] if idx + 9 < len(crows) else (None,)*4
            factor_cer = _sf(fcer_row[1])
            cers_row   = crows[idx + 10] if idx + 10 < len(crows) else (None,)*4
            cer_settl  = _sf(cers_row[1])
            cerb_row   = crows[idx + 11] if idx + 11 < len(crows) else (None,)*4
            cer_base   = _sf(cerb_row[1])
            bonos.append({
                'nombre':        nombre,
                'tipo_bono':     'Bonos CER',
                'ticker':        nombre,
                'tasa_cupon':    tasa_cupon,
                'base_calculo':  base_calculo,
                'periodicidad':  periodicidad,
                'fecha_emision': fecha_emision,
                'maturity':      maturity,
                'residual':      residual,
                'sink_factor':   sink_factor,
                'factor_cer':    factor_cer,
                'cer_settl':     cer_settl,
                'cer_base':      cer_base,
                'flujos':        [],
            })
            idx += CER_BLOCK

    if not bonos:
        st.error("❌ No se encontraron bonos en el archivo")
        st.stop()

    # Filtrar bonos vencidos
    _hoy = datetime.now().date()
    def _bono_vigente(b):
        # Lecaps y Bonos CER: usar campo maturity
        if b['tipo_bono'] in ('Lecaps & Boncaps', 'Bonos CER'):
            if b.get('maturity') is None:
                return True
            mat = b['maturity'].date() if hasattr(b['maturity'], 'date') else b['maturity']
            return mat >= _hoy
        # Resto: usar último flujo
        if not b['flujos']:
            return False
        ultima = max(
            f['fecha'].date() if hasattr(f['fecha'], 'date') else f['fecha']
            for f in b['flujos']
        )
        return ultima >= _hoy

    bonos = [b for b in bonos if _bono_vigente(b)]

    if not bonos:
        st.error("❌ No hay bonos vigentes disponibles")
        st.stop()

    # Generar tipos de bonos automáticamente a partir de los bonos procesados
    tipos_bono = sorted(set(b['tipo_bono'] for b in bonos))

    # Sidebar
    with st.sidebar:
        st.markdown("# CALCULADORA DE RENDIMIENTOS")

        # Filtro por tipo de bono
        if 'tipo_seleccionado' not in st.session_state:
            st.session_state.tipo_seleccionado = "Seleccione un Tipo"  # Valor inicial

        tipos_bono_con_seleccion = ["Seleccione un Tipo"] + tipos_bono

        # Usar una clave única que cambie cuando se presiona Volver
        if 'tipo_selectbox_key' not in st.session_state:
            st.session_state.tipo_selectbox_key = 0
        tipo_seleccionado = st.selectbox("Tipo de Bono", tipos_bono_con_seleccion, key=f"tipo_selectbox_{st.session_state.tipo_selectbox_key}")

        # Filtrar bonos por tipo
        if tipo_seleccionado == "Seleccione un Tipo":
            bonos_filtrados = bonos  # Mostrar todos los bonos cuando está en "Seleccione un Tipo"
        else:
            bonos_filtrados = [bono for bono in bonos if bono['tipo_bono'] == tipo_seleccionado]
        
        if not bonos_filtrados:
            st.error("No hay bonos del tipo seleccionado")
            st.stop()
        
        # Selección de bono
        nombres_bonos = [bono['nombre'] for bono in bonos_filtrados]
        nombres_bonos.sort()  # Ordenar alfabéticamente
        
        # Inicializar session_state si no existe
        if 'bono_seleccionado' not in st.session_state:
            st.session_state.bono_seleccionado = None
        if 'calcular' not in st.session_state:
            st.session_state.calcular = False
        bono_seleccionado = st.selectbox(
            "Elija un Bono", 
            nombres_bonos,
            index=None,  # Ningún bono seleccionado por defecto
            placeholder="Seleccione un bono...",
            key="bono_selectbox"
        )
        
        # Actualizar session_state cuando cambia la selección
        if bono_seleccionado != st.session_state.bono_seleccionado:
            st.session_state.bono_seleccionado = bono_seleccionado
            st.session_state.calcular = False
        
        # Variables para uso en la sección principal
        if bono_seleccionado and not st.session_state.get('flujos_bonos_seleccionados'):
            bono_actual = next((bono for bono in bonos_filtrados if bono['nombre'] == bono_seleccionado), None)
        else:
            st.session_state.calcular = False
            bono_actual = None
    
        # Calculadora de Flujos
        st.markdown("---")
        st.markdown("## CALCULADORA DE FLUJOS")
        
        # Inicializar session_state para flujos
        if 'flujos_calcular' not in st.session_state:
            st.session_state.flujos_calcular = False
        if 'flujos_bonos_seleccionados' not in st.session_state:
            st.session_state.flujos_bonos_seleccionados = []
        if 'flujos_tipo_selectbox_key' not in st.session_state:
            st.session_state.flujos_tipo_selectbox_key = 0
        if 'flujos_bono_selectbox_key' not in st.session_state:
            st.session_state.flujos_bono_selectbox_key = 0

        # Filtro por tipo de bono para flujos
        flujos_tipos_bono_con_seleccion = ["Seleccione un Tipo"] + tipos_bono
        flujos_tipo_seleccionado = st.selectbox(
            "Tipo de Bono",
            flujos_tipos_bono_con_seleccion,
            key=f"flujos_tipo_selectbox_{st.session_state.flujos_tipo_selectbox_key}"
        )

        # Filtrar bonos por tipo para flujos
        if flujos_tipo_seleccionado == "Seleccione un Tipo":
            flujos_bonos_filtrados = bonos
        else:
            flujos_bonos_filtrados = [b for b in bonos if b['tipo_bono'] == flujos_tipo_seleccionado]

        if not flujos_bonos_filtrados:
            st.error("No hay bonos del tipo seleccionado")
        else:
            flujos_nombres_bonos = sorted([b['nombre'] for b in flujos_bonos_filtrados])

            # Selectbox con key rotativa: se resetea a None cada vez que se agrega un bono
            bono_a_agregar = st.selectbox(
                "Agregar Bono",
                flujos_nombres_bonos,
                index=None,
                placeholder="Seleccione un bono...",
                key=f"flujos_bono_selectbox_{st.session_state.flujos_bono_selectbox_key}"
            )

            if bono_a_agregar:
                ya_en_lista = [b['nombre'] for b in st.session_state.flujos_bonos_seleccionados]
                if bono_a_agregar not in ya_en_lista:
                    bono_info = next((b for b in bonos if b['nombre'] == bono_a_agregar), None)
                    if bono_info:
                        ticker_flujo = bono_info.get('ticker', '').strip()
                        precio_api = None
                        if ticker_flujo and ticker_flujo != 'SPX500':
                            precio_api = obtener_precio_data912(ticker_flujo)
                        # Guardar precio con key por nombre
                        key_nombre = re.sub(r'[^a-zA-Z0-9]', '_', bono_a_agregar)
                        precio_key_flujo = f"precio_flujo_{key_nombre}"
                        st.session_state[precio_key_flujo] = float(precio_api) if precio_api and precio_api > 0 else 0.0
                        st.session_state.flujos_bonos_seleccionados.append({
                            'nombre': bono_a_agregar,
                            'nominales': '',
                            'precio': precio_api if precio_api and precio_api > 0 else '',
                            'info': bono_info
                        })
                # Siempre resetear el selectbox y limpiar selecciones de rendimientos
                st.session_state.flujos_bono_selectbox_key += 1
                st.session_state.flujos_calcular = False
                st.session_state.calcular = False
                st.session_state.bono_seleccionado = None
                for k in ('tipo_selectbox', 'bono_selectbox'):
                    if k in st.session_state:
                        del st.session_state[k]
                st.rerun()
        
        # Botón Volver para flujos (visible cuando hay bonos en la lista)
        if st.session_state.get('flujos_bonos_seleccionados'):
            if st.button("Volver", type="secondary", use_container_width=True, key="flujos_volver"):
                st.session_state.calcular = False
                st.session_state.bono_seleccionado = None
                st.session_state.flujos_calcular = False
                st.session_state.flujos_bonos_seleccionados = []
                st.session_state.tipo_selectbox_key += 1
                st.session_state.flujos_tipo_selectbox_key += 1
                st.session_state.flujos_bono_selectbox_key += 1
                for k in ('bono_selectbox', 'tipo_selectbox'):
                    if k in st.session_state:
                        del st.session_state[k]
                st.rerun()

        # Monitor
        st.markdown("---")
        if 'monitor' not in st.session_state:
            st.session_state.monitor = False
        if st.button("Monitor", use_container_width=True, key="btn_monitor"):
            st.session_state.monitor_prev_bono   = st.session_state.get('bono_seleccionado')
            st.session_state.monitor_prev_flujos = st.session_state.get('flujos_bonos_seleccionados', [])
            st.session_state.monitor = True
            st.session_state.bono_seleccionado = None
            st.session_state.flujos_bonos_seleccionados = []
            for _k in ('monitor_tick', 'monitor_panel'):
                st.session_state.pop(_k, None)
            st.rerun()


    # S2 (Calculadora de Flujos) - activo cuando hay bonos en la lista
    if st.session_state.get('flujos_bonos_seleccionados'):
        
        # Mostrar lista de bonos seleccionados con recuadro transparente
        st.markdown("""
        <div style="
            border: 1px solid #e0e0e0;
            border-radius: 8px;
            padding: 12px 16px;
            margin: 8px 0 20px 0;
            background-color: transparent;
            text-align: center;
            font-weight: 600;
            color: #333;
        ">
            Bonos Seleccionados para Flujos
        </div>
        """, unsafe_allow_html=True)
        
        if st.session_state.flujos_bonos_seleccionados:
            for i, bono in enumerate(st.session_state.flujos_bonos_seleccionados):
                # Keys basadas en nombre del bono (no índice) para evitar colisiones al borrar/reordenar
                key_nombre = re.sub(r'[^a-zA-Z0-9]', '_', bono['nombre'])
                precio_key = f"precio_flujo_{key_nombre}"
                nominales_key = f"nominales_{key_nombre}"

                # Layout con 6 columnas para incluir precio
                col1, col2, col3, col4, col5, col6 = st.columns([3, 1, 2, 1, 1, 1])

                with col1:
                    st.write(f"**{bono['nombre']}**")

                with col2:
                    st.write("Nominales:")

                with col3:
                    nominales_text = st.text_input(
                        "",
                        value=str(bono['nominales']) if bono['nominales'] != '' else "",
                        key=nominales_key,
                        label_visibility="collapsed"
                    )
                    try:
                        nominales = int(nominales_text) if nominales_text else 0
                    except ValueError:
                        nominales = 0
                    st.session_state.flujos_bonos_seleccionados[i]['nominales'] = nominales

                with col4:
                    st.write("Precio:")

                with col5:
                    # Inicializar precio en session_state si no existe aún
                    if precio_key not in st.session_state:
                        precio_actual = bono.get('precio', '')
                        if precio_actual and float(precio_actual) > 0:
                            st.session_state[precio_key] = float(precio_actual)
                        else:
                            st.session_state[precio_key] = 0.0

                    precio = st.number_input(
                        "",
                        min_value=0.0,
                        max_value=200.0,
                        step=0.01,
                        format="%.2f",
                        key=precio_key,
                        label_visibility="collapsed"
                    )
                    st.session_state.flujos_bonos_seleccionados[i]['precio'] = precio

                with col6:
                    if st.button("🗑️", key=f"remove_{i}_{key_nombre}", help="Eliminar bono"):
                        # Limpiar keys del bono eliminado
                        for k in [precio_key, nominales_key]:
                            if k in st.session_state:
                                del st.session_state[k]
                        st.session_state.flujos_bonos_seleccionados.pop(i)
                        st.session_state.flujos_calcular = False
                        st.rerun()
            
            # Botón para agregar más bonos
            st.markdown("---")
            
            # Botón para calcular flujos
            if st.button("Calcular Flujos", type="secondary", use_container_width=True):
                st.session_state.flujos_calcular = True
            
            # Mostrar tarjetas de métricas si se está calculando
            if st.session_state.get('flujos_calcular', False):
                # Obtener fecha actual
                fecha_actual = pd.Timestamp.now().date()
                
                st.markdown('<div class="metrics-grid">', unsafe_allow_html=True)
                
                # Primera fila de métricas
                col1, col2, col3, col4, col5 = st.columns(5)
                
                # Calcular métricas
                cantidad_bonos = len(st.session_state.flujos_bonos_seleccionados)
                total_intereses = 0
                total_amortizaciones = 0
                total_general = 0
                cupon_ponderado = 0.0
                total_nominales = 0
                suma_cupones_ponderados = 0
                
                # Calcular totales de los flujos y cupón ponderado
                if st.session_state.flujos_bonos_seleccionados:
                    for bono_item in st.session_state.flujos_bonos_seleccionados:
                        if bono_item['nominales'] > 0:
                            bono_info = bono_item['info']
                            nominales = bono_item['nominales']
                            
                            # Calcular cupón ponderado
                            # Detectar si el bono tiene cupón variable
                            cupones_unicos = set()
                            for flujo in bono_info['flujos']:
                                fecha_flujo = flujo['fecha']
                                if hasattr(fecha_flujo, 'date'):
                                    fecha_flujo = fecha_flujo.date()
                                
                                if fecha_flujo >= fecha_actual:
                                    cupones_unicos.add(flujo['cupon_vigente'])
                            
                            # Si hay múltiples cupones, calcular cupón ponderado del bono
                            if len(cupones_unicos) > 1:
                                # Bono con cupón variable - calcular cupón ponderado
                                suma_cupones_bono = 0
                                total_flujos_bono = 0
                                for flujo in bono_info['flujos']:
                                    fecha_flujo = flujo['fecha']
                                    if hasattr(fecha_flujo, 'date'):
                                        fecha_flujo = fecha_flujo.date()
                                    
                                    if fecha_flujo >= fecha_actual:
                                        suma_cupones_bono += flujo['cupon_vigente']
                                        total_flujos_bono += 1
                                
                                if total_flujos_bono > 0:
                                    cupon_ponderado_bono = suma_cupones_bono / total_flujos_bono
                                else:
                                    cupon_ponderado_bono = bono_info.get('tasa_cupon', 0)
                            else:
                                # Bono con cupón fijo
                                cupon_ponderado_bono = bono_info.get('tasa_cupon', 0)
                            
                            suma_cupones_ponderados += cupon_ponderado_bono * nominales
                            total_nominales += nominales
                            
                            for flujo in bono_info['flujos']:
                                fecha_cupon = flujo['fecha']
                                if hasattr(fecha_cupon, 'date'):
                                    fecha_cupon = fecha_cupon.date()
                                
                                if fecha_cupon >= fecha_actual:
                                    intereses = flujo['cupon'] * nominales / 100
                                    amortizaciones = flujo['capital'] * nominales / 100
                                    total = flujo['total'] * nominales / 100
                                    
                                    total_intereses += intereses
                                    total_amortizaciones += amortizaciones
                                    total_general += total
                
                # Calcular cupón ponderado final
                if total_nominales > 0:
                    cupon_ponderado = suma_cupones_ponderados / total_nominales
                
                with col1:
                    st.markdown(f'''
                    <div class="metric-card">
                        <div class="metric-label">Cantidad Bonos</div>
                        <div class="metric-value">{cantidad_bonos}</div>
                    </div>
                    ''', unsafe_allow_html=True)
                
                with col2:
                    st.markdown(f'''
                    <div class="metric-card">
                        <div class="metric-label">Total Intereses</div>
                        <div class="metric-value">${formatear_numero(total_intereses, 2)}</div>
                    </div>
                    ''', unsafe_allow_html=True)
                
                with col3:
                    st.markdown(f'''
                    <div class="metric-card">
                        <div class="metric-label">Total Amortizaciones</div>
                        <div class="metric-value">${formatear_numero(total_amortizaciones, 2)}</div>
                    </div>
                    ''', unsafe_allow_html=True)
                
                with col4:
                    st.markdown(f'''
                    <div class="metric-card">
                        <div class="metric-label">Total</div>
                        <div class="metric-value">${formatear_numero(total_general, 2)}</div>
                    </div>
                    ''', unsafe_allow_html=True)
                
                with col5:
                    # Calcular TIR solo si todos los bonos tienen precio
                    tir_calculada = None
                    todos_tienen_precio = True
                    
                    if st.session_state.flujos_bonos_seleccionados:
                        # Verificar que todos los bonos tengan precio
                        for bono_item in st.session_state.flujos_bonos_seleccionados:
                            if bono_item['nominales'] > 0 and (not bono_item.get('precio') or bono_item.get('precio', 0) <= 0):
                                todos_tienen_precio = False
                                break
                        
                        if todos_tienen_precio:
                            # Preparar datos para TIR: fechas y valores
                            fechas_tir = []
                            valores_tir = []
                            
                            # Agregar valor actual (inversión inicial)
                            # El precio se ingresa en base 100, se convierte a base 1 dividiendo por 100
                            total_inversion = 0
                            for bono_item in st.session_state.flujos_bonos_seleccionados:
                                if bono_item['nominales'] > 0 and bono_item.get('precio', 0) > 0:
                                    precio_base1 = bono_item['precio'] / 100  # Convertir de base 100 a base 1
                                    total_inversion += bono_item['nominales'] * precio_base1
                            
                            if total_inversion > 0:
                                fechas_tir.append(fecha_actual)
                                valores_tir.append(-total_inversion)  # Negativo (inversión)
                            
                            # Agregar flujos futuros
                            for bono_item in st.session_state.flujos_bonos_seleccionados:
                                if bono_item['nominales'] > 0:
                                    bono_info = bono_item['info']
                                    nominales = bono_item['nominales']
                                    
                                    for flujo in bono_info['flujos']:
                                        fecha_flujo = flujo['fecha']
                                        if hasattr(fecha_flujo, 'date'):
                                            fecha_flujo = fecha_flujo.date()
                                        
                                        if fecha_flujo >= fecha_actual:
                                            fechas_tir.append(fecha_flujo)
                                            total_flujo = flujo['total'] * nominales / 100
                                            valores_tir.append(total_flujo)  # Positivo (retorno)
                            
                            # Calcular TIR usando la misma función que la calculadora de rendimientos
                            if len(valores_tir) > 1:
                                try:
                                    # Preparar datos para calcular_ytm
                                    precio_dirty = -valores_tir[0]  # Valor actual (positivo)
                                    flujos = valores_tir[1:]  # Flujos futuros
                                    fechas = fechas_tir[1:]  # Fechas futuras
                                    
                                    # Usar la función calcular_ytm existente
                                    tir_efectiva = calcular_ytm(
                                        precio_dirty=precio_dirty,
                                        flujos=flujos,
                                        fechas=fechas,
                                        fecha_liquidacion=fecha_actual,
                                        base_calculo="ACT/365",
                                        periodicidad=2
                                    )
                                    
                                    # Convertir TIR efectiva a TIR semestral
                                    # Fórmula: TIR_semestral = 2 * ((1 + TIR_efectiva)^(1/2) - 1)
                                    tir_calculada = 2 * ((1 + tir_efectiva) ** (1/2) - 1)
                                    
                                except Exception as e:
                                    tir_calculada = 0.0
                    
                    # Mostrar TIR o valor vacío
                    if tir_calculada is not None:
                        st.markdown(f'''
                        <div class="metric-card">
                            <div class="metric-label">TIR consolidada</div>
                            <div class="metric-value">{tir_calculada:.2%}</div>
                        </div>
                        ''', unsafe_allow_html=True)
                    else:
                        st.markdown(f'''
                        <div class="metric-card">
                            <div class="metric-label">TIR consolidada</div>
                            <div class="metric-value">-</div>
                        </div>
                        ''', unsafe_allow_html=True)
                
                st.markdown('</div>', unsafe_allow_html=True)
            
            # Mostrar tabla de flujos si se presionó calcular
            if st.session_state.get('flujos_calcular', False):
                
                # Obtener fecha actual
                fecha_actual = pd.Timestamp.now().date()
                
                # Recopilar todos los flujos de todos los bonos seleccionados
                todos_flujos = []
                
                for bono_item in st.session_state.flujos_bonos_seleccionados:
                    if bono_item['nominales'] > 0:  # Solo si tiene nominales
                        bono_info = bono_item['info']
                        nominales = bono_item['nominales']
                        
                        # Calcular flujos para este bono (similar a S1)
                        flujos_bono = []
                        fechas_bono = []
                        
                        # Obtener fechas de cupones y amortizaciones
                        for flujo in bono_info['flujos']:
                            fecha_cupon = flujo['fecha']
                            if hasattr(fecha_cupon, 'date'):
                                fecha_cupon = fecha_cupon.date()
                            
                            if fecha_cupon >= fecha_actual:  # Solo fechas futuras
                                flujos_bono.append({
                                    'fecha': fecha_cupon,
                                    'activo': bono_item['nombre'],
                                    'cupon': flujo['cupon_vigente'],
                                    'intereses': round(flujo['cupon'] * nominales / 100, 2),
                                    'amortizaciones': round(flujo['capital'] * nominales / 100, 2),
                                    'total': round(flujo['total'] * nominales / 100, 2)
                                })
                                fechas_bono.append(fecha_cupon)
                        
                        todos_flujos.extend(flujos_bono)
                
                if todos_flujos:
                    # Calcular total de inversión (nominales × precio en base 1)
                    # El precio se ingresa en base 100, se convierte a base 1 dividiendo por 100
                    total_inversion = 0
                    for bono_item in st.session_state.flujos_bonos_seleccionados:
                        if bono_item['nominales'] > 0 and bono_item.get('precio', 0) > 0:
                            precio_base1 = bono_item['precio'] / 100  # Convertir de base 100 a base 1
                            total_inversion += bono_item['nominales'] * precio_base1
                    
                    # Agregar fila inicial con fecha actual y total de inversión
                    fila_inicial = {
                        'fecha': fecha_actual,
                        'activo': 'Valor Actual',
                        'cupon': 0.0,
                        'intereses': 0.0,
                        'amortizaciones': 0.0,
                        'total': total_inversion
                    }
                    
                    # Insertar fila inicial al principio
                    todos_flujos.insert(0, fila_inicial)
                    
                    # Ordenar por fecha
                    todos_flujos.sort(key=lambda x: x['fecha'])
                    
                    # Crear DataFrame
                    df_flujos = pd.DataFrame(todos_flujos)
                    
                    # Formatear fechas a DD/MM/YY
                    df_flujos['fecha'] = pd.to_datetime(df_flujos['fecha']).dt.strftime('%d/%m/%y')
                    
                    # Formatear números con separadores de miles
                    df_flujos['intereses'] = df_flujos['intereses'].apply(lambda x: formatear_numero(x, 2))
                    df_flujos['amortizaciones'] = df_flujos['amortizaciones'].apply(lambda x: formatear_numero(x, 2))
                    df_flujos['total'] = df_flujos['total'].apply(lambda x: formatear_numero(x, 2))
                    
                    # Formatear cupón como porcentaje con 2 decimales
                    df_flujos['cupon'] = df_flujos['cupon'].apply(lambda x: f"{x*100:.2f}%")
                    
                    # Mostrar tabla de flujos con el mismo estilo que la tabla inicial
                    df_flujos = df_flujos.rename(columns={
                        'fecha': 'Fecha',
                        'activo': 'Activo',
                        'cupon': 'Cupón',
                        'intereses': 'Intereses',
                        'amortizaciones': 'Amortizaciones',
                        'total': 'Total'
                    })
                    st.markdown(TABLA_BONOS_CSS, unsafe_allow_html=True)
                    st.markdown(
                        render_tabla_bonos_html(
                            df_flujos,
                            columnas_derecha=['Intereses', 'Amortizaciones', 'Total']
                        ),
                        unsafe_allow_html=True
                    )

                    # Gráfico de flujos por trimestre
                    st.markdown("---")
                    st.markdown("### Flujos por Trimestre (Próximos 5 Años)")

                    # Usar datos numéricos crudos (todos_flujos) antes del formateo de strings
                    df_futuros_raw = pd.DataFrame([
                        f for f in todos_flujos if f.get('activo') != 'Valor Actual'
                    ])

                    if not df_futuros_raw.empty:
                        df_futuros_raw['fecha'] = pd.to_datetime(df_futuros_raw['fecha'])
                        # Crear columna de trimestre
                        df_futuros_raw['año'] = df_futuros_raw['fecha'].dt.year
                        df_futuros_raw['trimestre'] = df_futuros_raw['fecha'].dt.quarter
                        df_futuros_raw['trimestre_label'] = df_futuros_raw['trimestre'].astype(str) + 'Q' + df_futuros_raw['año'].astype(str).str[2:]

                        # Filtrar solo los próximos 5 años
                        año_actual = pd.Timestamp.now().year
                        df_futuros_raw = df_futuros_raw[df_futuros_raw['año'] <= año_actual + 5]

                        # Agrupar por trimestre con valores ya numéricos
                        df_trimestral = df_futuros_raw.groupby(['año', 'trimestre', 'trimestre_label']).agg({
                            'intereses': 'sum',
                            'amortizaciones': 'sum'
                        }).reset_index()

                        # Crear etiqueta completa para el eje X
                        df_trimestral['periodo'] = df_trimestral['trimestre'].astype(str) + 'Q' + df_trimestral['año'].astype(str).str[2:]
                        df_trimestral = df_trimestral.sort_values(['año', 'trimestre'])
                        
                        if not df_trimestral.empty:
                            tiene_amort = df_trimestral['amortizaciones'].sum() > 0
                            tiene_int   = df_trimestral['intereses'].sum() > 0

                            _BG = '#f4f6fb'
                            _axis_style = dict(
                                showgrid=True, gridcolor='#cccccc',
                                linecolor='#999999', linewidth=1, showline=True,
                                tickfont=dict(color='#444444', size=11),
                                title_font=dict(color='#444444')
                            )

                            def _gradient_colors(values, base_rgb):
                                """Genera lista de colores con opacidad proporcional al valor."""
                                max_v = max(values) if max(values) > 0 else 1
                                r, g, b = base_rgb
                                return [
                                    f'rgba({r},{g},{b},{0.35 + 0.65 * v / max_v:.2f})'
                                    for v in values
                                ]

                            if tiene_amort and tiene_int:
                                colors_amort = _gradient_colors(df_trimestral['amortizaciones'], (74, 111, 165))
                                colors_int   = _gradient_colors(df_trimestral['intereses'],      (122, 179, 212))

                                fig = make_subplots(
                                    rows=2, cols=1,
                                    shared_xaxes=True,
                                    row_heights=[0.6, 0.4],
                                    vertical_spacing=0.10,
                                    subplot_titles=('Amortizaciones', 'Intereses')
                                )
                                fig.add_trace(go.Bar(
                                    name='Amortizaciones',
                                    x=df_trimestral['periodo'],
                                    y=df_trimestral['amortizaciones'],
                                    marker=dict(color=colors_amort, line_width=0),
                                    hovertemplate='<b>%{x}</b><br>Amortizaciones: $%{y:,.0f}<extra></extra>'
                                ), row=1, col=1)
                                fig.add_trace(go.Bar(
                                    name='Intereses',
                                    x=df_trimestral['periodo'],
                                    y=df_trimestral['intereses'],
                                    marker=dict(color=colors_int, line_width=0),
                                    hovertemplate='<b>%{x}</b><br>Intereses: $%{y:,.0f}<extra></extra>'
                                ), row=2, col=1)
                                fig.update_yaxes(tickformat='$,.0f', **_axis_style, row=1, col=1)
                                fig.update_yaxes(tickformat='$,.0f', **_axis_style, row=2, col=1)
                                fig.update_xaxes(**_axis_style, row=1, col=1)
                                fig.update_xaxes(title_text='Trimestre', **_axis_style, row=2, col=1)
                                fig.update_layout(
                                    height=560,
                                    plot_bgcolor=_BG, paper_bgcolor=_BG,
                                    showlegend=True,
                                    legend=dict(orientation='h', yanchor='bottom', y=1.02,
                                                xanchor='right', x=1,
                                                font=dict(color='#444444', size=12)),
                                    hovermode='x unified',
                                    bargap=0.2,
                                    margin=dict(t=60, b=40, l=60, r=20),
                                    font=dict(color='#444444')
                                )
                                for ann in fig.layout.annotations:
                                    ann.font.color = '#444444'
                                    ann.font.size  = 12
                            else:
                                # Solo un tipo de flujo
                                if tiene_amort:
                                    y_vals = df_trimestral['amortizaciones']
                                    colors = _gradient_colors(y_vals, (74, 111, 165))
                                    nombre = 'Amortizaciones'
                                else:
                                    y_vals = df_trimestral['intereses']
                                    colors = _gradient_colors(y_vals, (122, 179, 212))
                                    nombre = 'Intereses'
                                fig = go.Figure(go.Bar(
                                    name=nombre,
                                    x=df_trimestral['periodo'],
                                    y=y_vals,
                                    marker=dict(color=colors, line_width=0),
                                    hovertemplate=f'<b>%{{x}}</b><br>{nombre}: $%{{y:,.0f}}<extra></extra>'
                                ))
                                fig.update_layout(
                                    xaxis=dict(**_axis_style, title='Trimestre'),
                                    yaxis=dict(tickformat='$,.0f', **_axis_style),
                                    plot_bgcolor=_BG, paper_bgcolor=_BG,
                                    height=420, hovermode='x unified',
                                    bargap=0.2,
                                    margin=dict(t=40, b=40, l=60, r=20),
                                    font=dict(color='#444444')
                                )

                            st.plotly_chart(fig, use_container_width=True)
                        else:
                            st.info("No hay flujos futuros en los próximos 5 años")
                    else:
                        st.info("No hay flujos futuros para mostrar (ingresá nominales para ver el gráfico)")
                else:
                    st.warning("⚠️ No se encontraron flujos futuros para los bonos seleccionados")
        else:
            st.info("🔧 CALCULADORA DE FLUJOS - Pantalla lista para nuevas funcionalidades")
    
    # S1 (Calculadora de Rendimientos) - Mostrar cuando hay bono seleccionado (sin necesidad de calcular)
    elif st.session_state.bono_seleccionado and not st.session_state.get('flujos_bonos_seleccionados'):
        # Obtener el bono actual del session_state
        bono_actual = next((bono for bono in bonos_filtrados if bono['nombre'] == st.session_state.bono_seleccionado), None)
        if not bono_actual:
            st.session_state.bono_seleccionado = None
            st.rerun()

        # --- Lecaps & Boncaps: mismo layout, tarjetas en blanco por ahora ---
        if bono_actual.get('tipo_bono') == 'Lecaps & Boncaps':
            col1_lec, col2_lec = st.columns([1, 2])
            with col1_lec:
                # Fecha de liquidación
                st.markdown("<div class='inline-field-label'>Fecha de Liquidación</div>", unsafe_allow_html=True)
                st.date_input("", value=get_next_business_day(), format="DD/MM/YYYY",
                              key="fecha_liq_lecap", label_visibility="collapsed")
                st.markdown("<div class='inline-field-label'>Precio Dirty</div>", unsafe_allow_html=True)
                _key_lec = f"precio_lecap_{bono_actual['nombre']}"
                _precio_lec_default = st.session_state.get(_key_lec, 0.0) or 0.0
                _precio_manual_monitor_lec = obtener_precio_manual_monitor(bono_actual.get('tipo_bono'), bono_actual['nombre'])
                if _key_lec not in st.session_state:
                    if _precio_manual_monitor_lec is not None:
                        _precio_lec_default = float(_precio_manual_monitor_lec)
                        st.session_state[_key_lec] = _precio_lec_default
                    else:
                        try:
                            _pd_lec = {**obtener_precios_data912('arg_bonds'),
                                       **obtener_precios_data912('arg_corp'),
                                       **obtener_precios_data912('arg_notes')}
                            _tk_lec = bono_actual.get('ticker', '').strip().upper()
                            _pm_lec = _pd_lec.get(_tk_lec)
                            if _pm_lec and _pm_lec.get('c', 0) > 0:
                                _precio_lec_default = float(_pm_lec['c'])
                                st.session_state[_key_lec] = _precio_lec_default
                        except Exception:
                            pass
                st.number_input("", min_value=0.0, value=_precio_lec_default, step=0.01, format="%.2f",
                                key=_key_lec, label_visibility="collapsed")
                col_calc_lec, col_vol_lec = st.columns(2)
                with col_calc_lec:
                    st.button("Calcular", type="primary", use_container_width=True, key="calcular_lecap", disabled=True)
                with col_vol_lec:
                    if st.button("Volver", type="secondary", use_container_width=True, key="volver_lecap"):
                        st.session_state.bono_seleccionado = None
                        st.session_state.tipo_seleccionado = "Seleccione un Tipo"
                        st.session_state.calcular = False
                        st.session_state.tipo_selectbox_key += 1
                        st.session_state.flujos_tipo_selectbox_key = st.session_state.get('flujos_tipo_selectbox_key', 0) + 1
                        st.session_state.flujos_bono_selectbox_key = st.session_state.get('flujos_bono_selectbox_key', 0) + 1
                        for k in ('bono_selectbox', 'tipo_selectbox'):
                            if k in st.session_state:
                                del st.session_state[k]
                        st.rerun()
                # Información del Bono
                mat = bono_actual.get('maturity')
                mat_str = mat.strftime('%d/%m/%Y') if mat and hasattr(mat, 'strftime') else 'N/A'
                st.markdown(f"""
                <div class="calc-card-fill">
                    <div class="calc-card-title">Información del Bono</div>
                    <div style="font-size:0.92rem; color:#444; line-height:1.98;">
                        <p style="margin:0.37rem 0;"><strong style="color:#1a237e;">Nombre:</strong> {bono_actual['nombre']}</p>
                        <p style="margin:0.37rem 0;"><strong style="color:#1a237e;">Vencimiento:</strong> {mat_str}</p>
                        <p style="margin:0.37rem 0;"><strong style="color:#1a237e;">Tasa de cupón:</strong> {bono_actual['tasa_cupon']:.2%}</p>
                        <p style="margin:0.37rem 0;"><strong style="color:#1a237e;">Periodicidad:</strong> -</p>
                        <p style="margin:0.37rem 0;"><strong style="color:#1a237e;">Base de cálculo:</strong> 30/360</p>
                        <p style="margin:0.37rem 0;"><strong style="color:#1a237e;">Ticker:</strong> {bono_actual['ticker']}</p>
                    </div>
                </div>
                """, unsafe_allow_html=True)
            with col2_lec:
                _precio_lec = st.session_state.get(f"precio_lecap_{bono_actual['nombre']}", 0.0) or 0.0
                _vf_lec = bono_actual.get('valor_final', 0) or 0

                # Días remanente: vencimiento − fecha liquidación
                _fecha_liq_lec = st.session_state.get('fecha_liq_lecap', get_next_business_day())
                _mat_lec = bono_actual.get('maturity')
                if _mat_lec and _fecha_liq_lec:
                    _mat_date_lec = _mat_lec.date() if hasattr(_mat_lec, 'date') else _mat_lec
                    _dr_lec = max((_mat_date_lec - _fecha_liq_lec).days, 0)
                else:
                    _dr_lec = 0

                # Vida Media (base anual, estructura bullet): único flujo al vencimiento
                _vm_lec = _dr_lec / 365.0 if _dr_lec > 0 else 0.0  # en años

                # Macaulay Duration = Vida Media para zero-coupon bullet
                _macaulay_lec = _vm_lec

                # TNA (interés simple): (VF − P) / P / t * 365
                _tna_lec = (_vf_lec - _precio_lec) / _precio_lec / _dr_lec * 365 if _precio_lec > 0 and _dr_lec > 0 else None

                # TEA (interés compuesto): (1 + (VF-P)/P) ^ (365/días) - 1
                _tea_lec = (1 + (_vf_lec - _precio_lec) / _precio_lec) ** (365.0 / _dr_lec) - 1 if _precio_lec > 0 and _dr_lec > 0 else None

                # TEM: (1 + TEA) ^ (1/12) - 1
                _tem_lec = (1 + _tea_lec) ** (1 / 12) - 1 if _tea_lec is not None else None

                # Modified Duration = Macaulay / (1 + TEA)  [interés compuesto]
                _mod_dur_lec = _macaulay_lec / (1 + _tea_lec) if _tea_lec is not None and _vm_lec > 0 else None

                st.markdown(f'''
                <div class="metrics-card">
                    <div class="metrics-card-title">Rendimiento y duración</div>
                    <div class="metrics-row">
                        <div class="metric-card"><div class="metric-label">TNA</div><div class="metric-value">{f"{_tna_lec:.4%}" if _tna_lec is not None else "-"}</div></div>
                        <div class="metric-card"><div class="metric-label">TEM</div><div class="metric-value">{f"{_tem_lec:.4%}" if _tem_lec is not None else "-"}</div></div>
                        <div class="metric-card"><div class="metric-label">Duración Modificada</div><div class="metric-value">{formatear_numero(_mod_dur_lec, 2) if _mod_dur_lec is not None else "-"}</div></div>
                        <div class="metric-card"><div class="metric-label">Valor Final</div><div class="metric-value">{formatear_numero(_vf_lec, 4)}</div></div>
                    </div>
                </div>
                <div class="metrics-card">
                    <div class="metrics-card-title">Otros indicadores</div>
                    <div class="metrics-row">
                        <div class="metric-card"><div class="metric-label">Días Remanente</div><div class="metric-value">{int(_dr_lec) if _dr_lec > 0 else "-"}</div></div>
                        <div class="metric-card"><div class="metric-label">Vida Media</div><div class="metric-value">{formatear_numero(_vm_lec, 2) if _vm_lec > 0 else "-"}</div></div>
                        <div class="metric-card"><div class="metric-label">Macaulay Duration</div><div class="metric-value">{formatear_numero(_macaulay_lec, 2) if _macaulay_lec > 0 else "-"}</div></div>
                        <div class="metric-card"><div class="metric-label"></div><div class="metric-value"></div></div>
                    </div>
                </div>
                ''', unsafe_allow_html=True)
            st.stop()

        # --- Bonos CER: mismo layout que Lecaps ---
        if bono_actual.get('tipo_bono') == 'Bonos CER':
            col1_cer, col2_cer = st.columns([1, 2])
            with col1_cer:
                st.markdown("<div class='inline-field-label'>Fecha de Liquidación</div>", unsafe_allow_html=True)
                st.date_input("", value=get_next_business_day(), format="DD/MM/YYYY",
                              key="fecha_liq_cer", label_visibility="collapsed")
                st.markdown("<div class='inline-field-label'>Precio Dirty</div>", unsafe_allow_html=True)
                _precio_cer_default = st.session_state.get(f"precio_cer_{bono_actual['nombre']}", 0.0) or 0.0
                _precio_manual_monitor_cer = obtener_precio_manual_monitor(bono_actual.get('tipo_bono'), bono_actual['nombre'])
                if f"precio_cer_{bono_actual['nombre']}" not in st.session_state:
                    if _precio_manual_monitor_cer is not None:
                        _precio_cer_default = float(_precio_manual_monitor_cer)
                        st.session_state[f"precio_cer_{bono_actual['nombre']}"] = _precio_cer_default
                    else:
                        # Intentar obtener precio de la API si no viene de la tabla
                        try:
                            import requests as _rcer
                            _pd = {**obtener_precios_data912('arg_bonds'), **obtener_precios_data912('arg_notes')}
                            _pm = _pd.get(bono_actual['ticker'].upper())
                            if _pm and _pm['c'] > 0:
                                _precio_cer_default = float(_pm['c'])
                                st.session_state[f"precio_cer_{bono_actual['nombre']}"] = _precio_cer_default
                        except Exception:
                            pass
                st.number_input("", min_value=0.0, value=_precio_cer_default, step=0.01, format="%.2f",
                                key=f"precio_cer_{bono_actual['nombre']}", label_visibility="collapsed")
                col_calc_cer, col_vol_cer = st.columns(2)
                with col_calc_cer:
                    st.button("Calcular", type="primary", use_container_width=True, key="calcular_cer", disabled=True)
                with col_vol_cer:
                    if st.button("Volver", type="secondary", use_container_width=True, key="volver_cer"):
                        st.session_state.bono_seleccionado = None
                        st.session_state.tipo_seleccionado = "Seleccione un Tipo"
                        st.session_state.calcular = False
                        st.session_state.tipo_selectbox_key += 1
                        st.session_state.flujos_tipo_selectbox_key = st.session_state.get('flujos_tipo_selectbox_key', 0) + 1
                        st.session_state.flujos_bono_selectbox_key = st.session_state.get('flujos_bono_selectbox_key', 0) + 1
                        for k in ('bono_selectbox', 'tipo_selectbox'):
                            if k in st.session_state:
                                del st.session_state[k]
                        st.rerun()
                mat_cer = bono_actual.get('maturity')
                mat_cer_str = mat_cer.strftime('%d/%m/%Y') if mat_cer and hasattr(mat_cer, 'strftime') else 'N/A'
                st.markdown(f"""
                <div class="calc-card-fill">
                    <div class="calc-card-title">Información del Bono</div>
                    <div style="font-size:0.92rem; color:#444; line-height:1.98;">
                        <p style="margin:0.37rem 0;"><strong style="color:#1a237e;">Nombre:</strong> {bono_actual['nombre']}</p>
                        <p style="margin:0.37rem 0;"><strong style="color:#1a237e;">Vencimiento:</strong> {mat_cer_str}</p>
                        <p style="margin:0.37rem 0;"><strong style="color:#1a237e;">Tasa de cupón:</strong> {bono_actual['tasa_cupon']:.2%}</p>
                        <p style="margin:0.37rem 0;"><strong style="color:#1a237e;">Periodicidad:</strong> {bono_actual.get('periodicidad', '-')}</p>
                        <p style="margin:0.37rem 0;"><strong style="color:#1a237e;">Base de cálculo:</strong> {bono_actual.get('base_calculo', '30/360')}</p>
                        <p style="margin:0.37rem 0;"><strong style="color:#1a237e;">Ticker:</strong> {bono_actual['ticker']}</p>
                    </div>
                </div>
                """, unsafe_allow_html=True)
            with col2_cer:
                # Obtener CER de liquidación desde BCRA (10° día hábil anterior)
                _fecha_liq_cer = st.session_state.get('fecha_liq_cer', get_next_business_day())
                _cer_settlement, _cer_settlement_fecha = obtener_cer_settlement(_fecha_liq_cer)
                if _cer_settlement:
                    _cer_settl_str = (
                        f"{formatear_numero(_cer_settlement, 4)}"
                        f"<br><span style='font-size:0.7rem;color:#888'>({_cer_settlement_fecha.strftime('%d/%m/%Y') if _cer_settlement_fecha else ''})</span>"
                    )
                else:
                    _cer_fecha_req = n_dias_habiles_antes(_fecha_liq_cer, 10) if _fecha_liq_cer else None
                    _cer_settl_str = (
                        f"-<br><span style='font-size:0.7rem;color:#e65100;'>⚠ Sin dato para "
                        f"{_cer_fecha_req.strftime('%d/%m/%Y') if _cer_fecha_req else '?'}</span>"
                    )

                # Cálculos para TIR Real
                _precio_cer = st.session_state.get(f"precio_cer_{bono_actual['nombre']}", 0.0) or 0.0
                _mat_cer2 = bono_actual.get('maturity')
                if _mat_cer2 and _fecha_liq_cer:
                    _mat_date_cer = _mat_cer2.date() if hasattr(_mat_cer2, 'date') else _mat_cer2
                    _liq_date_cer = _fecha_liq_cer.date() if hasattr(_fecha_liq_cer, 'date') else _fecha_liq_cer
                    _dr_cer = max((_mat_date_cer - _liq_date_cer).days, 0)
                else:
                    _dr_cer = 0
                _cer_base_val = bono_actual.get('cer_base') or 0
                _factor_cer_vivo = (_cer_settlement / _cer_base_val) if (_cer_settlement and _cer_base_val) else None
                # TIR Real = XIRR(-precio, factor_cer*100; fecha_liq, fecha_vto)
                # Con 2 flujos: (factor_cer*100 / precio)^(365/días) - 1
                if _precio_cer > 0 and _dr_cer > 0 and _factor_cer_vivo:
                    _tir_real_cer = (_factor_cer_vivo * 100 / _precio_cer) ** (365.0 / _dr_cer) - 1
                else:
                    _tir_real_cer = None

                # Vida Media y Duración Modificada (bullet, datos disponibles)
                _vm_cer = _dr_cer / 365.0 if _dr_cer > 0 else None
                _dur_mod_cer = _vm_cer / (1 + _tir_real_cer) if (_vm_cer and _tir_real_cer is not None) else None

                st.markdown(f'''
                <div class="metrics-card">
                    <div class="metrics-card-title">Rendimiento y duración</div>
                    <div class="metrics-row">
                        <div class="metric-card"><div class="metric-label">TIR Efectiva Anual</div><div class="metric-value">{f"{_tir_real_cer:.4%}" if _tir_real_cer is not None else "-"}</div></div>
                        <div class="metric-card"><div class="metric-label">TEM</div><div class="metric-value">{f"{((1 + _tir_real_cer) ** (30/360) - 1):.4%}" if _tir_real_cer is not None else "-"}</div></div>
                        <div class="metric-card"><div class="metric-label">Duración Modificada</div><div class="metric-value">{formatear_numero(_dur_mod_cer, 2) if _dur_mod_cer is not None else "-"}</div></div>
                        <div class="metric-card"><div class="metric-label">Vida Media</div><div class="metric-value">{formatear_numero(_vm_cer, 2) if _vm_cer is not None else "-"}</div></div>
                    </div>
                </div>
                <div class="metrics-card">
                    <div class="metrics-card-title">Otros indicadores</div>
                    <div class="metrics-row">
                        <div class="metric-card"><div class="metric-label">CER Base</div><div class="metric-value">{formatear_numero(bono_actual.get("cer_base", 0), 4)}</div></div>
                        <div class="metric-card"><div class="metric-label">CER Settlement</div><div class="metric-value">{_cer_settl_str}</div></div>
                        <div class="metric-card"><div class="metric-label">Factor CER</div><div class="metric-value">{formatear_numero(_cer_settlement / bono_actual.get("cer_base", 1) if _cer_settlement and bono_actual.get("cer_base") else None, 4) if _cer_settlement and bono_actual.get("cer_base") else "-"}</div></div>
                        <div class="metric-card"><div class="metric-label"></div><div class="metric-value"></div></div>
                    </div>
                </div>
                ''', unsafe_allow_html=True)
            st.stop()

        # Layout principal
        col1, col2 = st.columns([1, 2])
        
        # COLUMNA IZQUIERDA - INPUTS Y BOTONES
        with col1:
            # Solo mostrar inputs si hay bono seleccionado Y NO hay bonos en S2
            if st.session_state.bono_seleccionado and not st.session_state.get('flujos_bonos_seleccionados'):
                # Encontrar el bono seleccionado
                bono_actual_main = next((bono for bono in bonos_filtrados if bono['nombre'] == st.session_state.bono_seleccionado), None)
                if not bono_actual_main:
                    st.session_state.bono_seleccionado = None
                    st.rerun()
                
                # Keys únicas por bono para sincronizar precios en USD/ARS
                precio_key_main = f"precio_dirty_usd_{st.session_state.bono_seleccionado}"
                precio_pesos_key_main = f"precio_dirty_pesos_{st.session_state.bono_seleccionado}"
                precio_prev_usd_key = f"precio_dirty_prev_usd_{st.session_state.bono_seleccionado}"
                precio_prev_pesos_key = f"precio_dirty_prev_pesos_{st.session_state.bono_seleccionado}"
                precio_fx_prev_key = f"precio_dirty_prev_fx_{st.session_state.bono_seleccionado}"
                precio_last_edited_key = f"precio_dirty_last_edited_{st.session_state.bono_seleccionado}"
                precio_init_key = f"precio_dirty_init_{st.session_state.bono_seleccionado}"

                # Inicializar el precio base en USD desde data912 una sola vez por bono
                if precio_init_key not in st.session_state:
                    ticker = bono_actual_main.get('ticker', '').strip()
                    precio_inicial = None
                    precio_manual_monitor = obtener_precio_manual_monitor(
                        bono_actual_main.get('tipo_bono'),
                        bono_actual_main['nombre']
                    )

                    if precio_manual_monitor is not None:
                        precio_inicial = float(precio_manual_monitor)
                    elif ticker and ticker != '' and ticker != 'SPX500':
                        precio_inicial = obtener_precio_data912(ticker)

                    st.session_state[precio_key_main] = float(precio_inicial) if precio_inicial and precio_inicial > 0 else 0.0
                    st.session_state[precio_pesos_key_main] = 0.0
                    st.session_state[precio_prev_usd_key] = st.session_state[precio_key_main]
                    st.session_state[precio_prev_pesos_key] = st.session_state[precio_pesos_key_main]
                    st.session_state[precio_fx_prev_key] = None
                    st.session_state[precio_last_edited_key] = 'usd'
                    st.session_state[precio_init_key] = True
                    st.session_state.calcular = True
                
                st.markdown("""
                <style>
                .inline-field-label {
                    font-size: 0.95rem;
                    font-weight: 400;
                    color: rgb(49, 51, 63);
                    margin-top: 0.35rem;
                    margin-bottom: 0.25rem;
                    line-height: 1.4;
                }
                div[data-testid="stSelectbox"] div[role="combobox"] {
                    font-size: 0.85rem !important;
                }
                div[data-testid="stDateInput"] input,
                div[data-testid="stNumberInput"] input,
                div[data-testid="stTextInput"] input {
                    font-size: 0.95rem !important;
                    font-weight: 500 !important;
                }
                div[data-testid="stTextInput"] {
                    margin-top: 0 !important;
                }
                div[data-testid="stTextInput"] > div {
                    margin-top: 0 !important;
                }
                div[data-testid="stSelectbox"] label[data-testid="stWidgetLabel"],
                div[data-testid="stDateInput"] label[data-testid="stWidgetLabel"],
                div[data-testid="stNumberInput"] label[data-testid="stWidgetLabel"],
                div[data-testid="stTextInput"] label[data-testid="stWidgetLabel"] {
                    display: none !important;
                    height: 0 !important;
                    margin: 0 !important;
                    padding: 0 !important;
                }
                </style>
                """, unsafe_allow_html=True)

                col_label_fecha, col_label_mep = st.columns(2)
                with col_label_fecha:
                    st.markdown("<div class='inline-field-label'>Fecha de Liquidación</div>", unsafe_allow_html=True)
                with col_label_mep:
                    tipo_cambio_tipo = st.selectbox(
                        "",
                        options=["Tipo de Cambio MEP", "Tipo de Cambio CCL"],
                        index=0,
                        key="tipo_cambio_tipo_main",
                        label_visibility="collapsed"
                    )

                tipo_cambio_valor = obtener_tipo_cambio_implicito_data912(tipo_cambio_tipo)
                tipo_cambio_valor_str = formatear_numero(tipo_cambio_valor, 2) if tipo_cambio_valor is not None else ""

                # Sincronización bidireccional USD <-> Pesos
                fx_actual = float(tipo_cambio_valor) if tipo_cambio_valor is not None else None
                usd_actual = float(st.session_state.get(precio_key_main, 0.0) or 0.0)
                pesos_actual = float(st.session_state.get(precio_pesos_key_main, 0.0) or 0.0)
                usd_prev = float(st.session_state.get(precio_prev_usd_key, usd_actual) or 0.0)
                pesos_prev = float(st.session_state.get(precio_prev_pesos_key, pesos_actual) or 0.0)
                fx_prev = st.session_state.get(precio_fx_prev_key)
                last_edited = st.session_state.get(precio_last_edited_key, 'usd')

                if fx_actual and fx_actual > 0:
                    if fx_prev is None:
                        pesos_actual = usd_actual * fx_actual
                    else:
                        usd_changed = abs(usd_actual - usd_prev) > 1e-9
                        pesos_changed = abs(pesos_actual - pesos_prev) > 1e-9
                        fx_changed = abs(float(fx_prev) - fx_actual) > 1e-9

                        if fx_changed:
                            if last_edited == 'pesos':
                                usd_actual = pesos_actual / fx_actual if fx_actual else 0.0
                            else:
                                pesos_actual = usd_actual * fx_actual
                        elif usd_changed and not pesos_changed:
                            pesos_actual = usd_actual * fx_actual
                            last_edited = 'usd'
                        elif pesos_changed and not usd_changed:
                            usd_actual = pesos_actual / fx_actual if fx_actual else 0.0
                            last_edited = 'pesos'

                    st.session_state[precio_key_main] = round(usd_actual, 2)
                    st.session_state[precio_pesos_key_main] = round(pesos_actual, 2)
                    st.session_state[precio_prev_usd_key] = st.session_state[precio_key_main]
                    st.session_state[precio_prev_pesos_key] = st.session_state[precio_pesos_key_main]
                    st.session_state[precio_fx_prev_key] = fx_actual
                    st.session_state[precio_last_edited_key] = last_edited

                col_fecha, col_mep = st.columns(2)
                with col_fecha:
                    fecha_liquidacion = st.date_input(
                        "",
                        value=get_next_business_day(),
                        format="DD/MM/YYYY",
                        key="fecha_liquidacion_main",
                        label_visibility="collapsed"
                    )
                with col_mep:
                    st.text_input(
                        "",
                        key=f"tipo_cambio_mep_main_{tipo_cambio_tipo}",
                        value=tipo_cambio_valor_str,
                        placeholder="",
                        label_visibility="collapsed",
                        disabled=True
                    )

                col_label_precio, col_label_precio_pesos = st.columns(2)
                with col_label_precio:
                    st.markdown("<div class='inline-field-label'>Precio Dirty en USD</div>", unsafe_allow_html=True)
                with col_label_precio_pesos:
                    st.markdown("<div class='inline-field-label'>Precio Dirty en Pesos</div>", unsafe_allow_html=True)

                col_inputs, col_input_precio_pesos = st.columns(2)
                with col_inputs:
                    precio_dirty = st.number_input(
                        "",
                        min_value=0.0,
                        max_value=200.0,
                        step=0.01,
                        format="%.2f",
                        key=precio_key_main,
                        help="El precio se obtiene automáticamente desde data912.com. Podés modificarlo manualmente.",
                        label_visibility="collapsed",
                        on_change=lambda: st.session_state.update({"calcular": True})
                    )

                with col_input_precio_pesos:
                    st.number_input(
                        "",
                        min_value=0.0,
                        step=0.01,
                        format="%.2f",
                        key=precio_pesos_key_main,
                        label_visibility="collapsed"
                    )

                col_calc, col_volver = st.columns(2)
                with col_calc:
                    if st.button("Calcular", type="primary", use_container_width=True, key="calcular_main"):
                        st.session_state.calcular = True
                with col_volver:
                    if st.button("Volver", type="secondary", use_container_width=True, key="volver_main"):
                            # Resetear TODAS las selecciones - estado inicial completo
                            st.session_state.calcular = False
                            st.session_state.bono_seleccionado = None
                            st.session_state.tipo_seleccionado = "Seleccione un Tipo"
                            
                            # Limpiar keys de precio
                            keys_to_delete = [key for key in st.session_state.keys() if key.startswith('precio_dirty_')]
                            for key in keys_to_delete:
                                del st.session_state[key]
                            
                            # Limpiar TODAS las selecciones de flujos
                            st.session_state.flujos_calcular = False
                            st.session_state.flujos_bonos_seleccionados = []
                            st.session_state.tipo_selectbox_key += 1
                            st.session_state.flujos_tipo_selectbox_key += 1
                            st.session_state.flujos_bono_selectbox_key += 1
                            for k in ('bono_selectbox', 'tipo_selectbox'):
                                if k in st.session_state:
                                    del st.session_state[k]
                            st.rerun()
                
                # Mapear periodicidad a texto
                periodicidad_texto_info = {
                    1: "anual",
                    2: "semestral",
                    3: "trimestral",
                    4: "trimestral",
                    6: "bimestral",
                    12: "mensual"
                }.get(bono_actual_main['periodicidad'], f"{bono_actual_main['periodicidad']} veces al año")

                fecha_vencimiento_info = encontrar_fecha_vencimiento(bono_actual_main['flujos'])
                from datetime import date
                cupon_vigente_actual_info = encontrar_cupon_vigente(date.today(), bono_actual_main['flujos'])

                st.markdown(f"""
                <div class="calc-card-fill">
                    <div class="calc-card-title">Información del Bono</div>
                    <div style="font-size:0.92rem; color:#444; line-height:1.98;">
                        <p style="margin:0.37rem 0;"><strong style="color:#1a237e;">Nombre:</strong> {bono_actual_main['nombre']}</p>
                        <p style="margin:0.37rem 0;"><strong style="color:#1a237e;">Vencimiento:</strong> {fecha_vencimiento_info.strftime('%d/%m/%Y') if fecha_vencimiento_info else 'N/A'}</p>
                        <p style="margin:0.37rem 0;"><strong style="color:#1a237e;">Tasa de cupón:</strong> {cupon_vigente_actual_info:.2%}</p>
                        <p style="margin:0.37rem 0;"><strong style="color:#1a237e;">Periodicidad:</strong> {periodicidad_texto_info}</p>
                        <p style="margin:0.37rem 0;"><strong style="color:#1a237e;">Base de cálculo:</strong> {bono_actual_main['base_calculo']}</p>
                        <p style="margin:0.37rem 0;"><strong style="color:#1a237e;">Ticker:</strong> {bono_actual_main['ticker']}</p>
                    </div>
                </div>
                """, unsafe_allow_html=True)
        
        # Espaciado adicional antes de Flujo de Fondos
        if st.session_state.calcular:
            st.markdown("<br><br>", unsafe_allow_html=True)
        
        # CÁLCULOS Y FLUJO DE FONDOS (solo si se presionó Calcular)
        if st.session_state.calcular:
            # Obtener fecha_liquidacion y precio_dirty desde session_state
            # Usar la misma key que se usó para guardar el precio
            precio_key_main = f"precio_dirty_usd_{st.session_state.bono_seleccionado}"
            fecha_liquidacion = st.session_state.get('fecha_liquidacion_main', get_next_business_day())
            precio_dirty = st.session_state.get(precio_key_main, 100.0)
            
            # Convertir fecha_liquidacion a datetime para comparación
            fecha_liquidacion_dt = pd.to_datetime(fecha_liquidacion)
            
            # Calcular flujos de caja
            flujos = []
            fechas = []
            flujos_capital = []
            
            for flujo in bono_actual['flujos']:
                if flujo['fecha'] > fecha_liquidacion_dt:
                    flujos.append(flujo['total'])
                    fechas.append(flujo['fecha'])
                    flujos_capital.append(flujo['capital'])
            
            if not flujos:
                st.error("❌ No hay flujos futuros para calcular")
            else:
                # SECCIÓN FLUJO DE FONDOS - FORMATO MEJORADO (ocupa todo el ancho, después de Información del Bono)
                st.markdown("## Flujo de Fondos")
                
                # Crear DataFrame con formato mejorado
                # Primera fila: fecha de liquidación y precio pagado (negativo)
                fechas_con_liquidacion = [fecha_liquidacion] + fechas
                capital_con_liquidacion = [""] + [formatear_numero(c, 1) if c > 0 else "" for c in flujos_capital]
                cupon_con_liquidacion = [""] + [formatear_numero(f-c, 1) for f, c in zip(flujos, flujos_capital)]
                total_con_liquidacion = [f"-{formatear_numero(precio_dirty, 1)}"] + [formatear_numero(f, 1) for f in flujos]
                
                df_simple = pd.DataFrame({
                    'Fecha': [f.strftime('%d/%m/%Y') for f in fechas_con_liquidacion],
                    'Capital': capital_con_liquidacion,
                    'Cupón': cupon_con_liquidacion,
                    'Total': total_con_liquidacion
                })
                
                # Mostrar tabla con el mismo estilo que la tabla inicial
                st.markdown(TABLA_BONOS_CSS, unsafe_allow_html=True)
                st.markdown(
                    render_tabla_bonos_html(
                        df_simple,
                        columnas_derecha=['Capital', 'Cupón', 'Total']
                    ),
                    unsafe_allow_html=True
                )
                
                # Gráfico del bono seleccionado - Solo si el ticker está disponible y es válido
                ticker_bono = bono_actual.get('ticker', '').strip()
                # Validar que el ticker existe, no está vacío, y no es un símbolo especial
                if ticker_bono and ticker_bono != '' and ticker_bono != 'SPX500':
                    bono_avanzado_html = f"""
                    <div class="tradingview-widget-container" style="height: 500px; width: 100%;">
                        <div class="tradingview-widget-container__widget" style="height: 100%; width: 100%;"></div>
                        <script type="text/javascript" src="https://s3.tradingview.com/external-embedding/embed-widget-mini-symbol-overview.js" async>
                        {{
                        "symbol": "{ticker_bono}",
                        "width": "100%",
                        "height": "500",
                        "locale": "es",
                        "dateRange": "12M",
                        "colorTheme": "light",
                        "isTransparent": true,
                        "autosize": false,
                        "largeChartUrl": "",
                        "hideTopToolbar": true,
                        "hideLegend": false,
                        "saveImage": false
                        }}
                        </script>
                    </div>
                    """
                    st.components.v1.html(bono_avanzado_html, height=500)
                
                # JavaScript para prevenir scroll automático al hacer clic en Calcular
                st.markdown("""
                <script>
                (function() {
                    // Scroll hacia arriba al hacer clic en Calcular (sin prevenir el comportamiento por defecto)
                    document.addEventListener('click', function(e) {
                        if (e.target.matches('button[data-testid="baseButton-primary"]') ||
                            e.target.textContent.trim() === 'Calcular') {
                            setTimeout(function() {
                                window.scrollTo({ top: 0, left: 0, behavior: 'auto' });
                            }, 200);
                        }
                    });
                })();
                </script>
                """, unsafe_allow_html=True)
                
                # COLUMNA DERECHA - RESULTADOS (métricas)
                with col2:
                    # Continuar con los cálculos solo si hay flujos
                    # Calcular YTM
                    ytm_efectiva = calcular_ytm(
                        precio_dirty,
                        flujos,
                        fechas,
                        fecha_liquidacion,
                        bono_actual['base_calculo'],
                        bono_actual['periodicidad']
                    )
                
                    # Calcular YTM anualizada según periodicidad
                    ytm_anualizada = bono_actual['periodicidad'] * ((1 + ytm_efectiva) ** (1 / bono_actual['periodicidad']) - 1)
                
                    # Calcular duración Macaulay
                    duracion_macaulay = calcular_duracion_macaulay(
                        flujos,
                        fechas,
                        fecha_liquidacion,
                        ytm_efectiva,
                        bono_actual['base_calculo']
                    )
                
                    # Calcular duración modificada
                    duracion_modificada = calcular_duracion_modificada(
                        duracion_macaulay,
                        ytm_anualizada / bono_actual['periodicidad'],
                        bono_actual['periodicidad']
                    )
                
                    # Calcular capital residual
                    capital_residual = 100 - sum([flujo['capital'] for flujo in bono_actual['flujos'] if flujo['fecha'] <= fecha_liquidacion_dt])
                
                    # Calcular intereses corridos
                    fecha_ultimo_cupon = encontrar_ultimo_cupon(fecha_liquidacion_dt, [flujo['fecha'] for flujo in bono_actual['flujos']], bono_actual.get('fecha_emision'))
                    if fecha_ultimo_cupon:
                        intereses_corridos = calcular_intereses_corridos(
                            fecha_liquidacion,
                            fecha_ultimo_cupon,
                            bono_actual['tasa_cupon'],
                            capital_residual,
                            bono_actual['base_calculo']
                        )
                    else:
                        intereses_corridos = 0
                
                    # Calcular precio limpio
                    precio_limpio = precio_dirty - intereses_corridos
                
                    # Calcular vida media
                    vida_media = calcular_vida_media(
                        flujos_capital,
                        fechas,
                        fecha_liquidacion,
                        bono_actual['base_calculo']
                    )
                
                    # Calcular paridad
                    valor_tecnico = capital_residual + intereses_corridos
                    paridad = precio_limpio / valor_tecnico if valor_tecnico > 0 else 0
                
                    # Encontrar próximo cupón
                    proximo_cupon = encontrar_proximo_cupon(fecha_liquidacion_dt, [flujo['fecha'] for flujo in bono_actual['flujos']])
                
                    # Calcular cupón vigente
                    cupon_vigente = encontrar_cupon_vigente(fecha_liquidacion, bono_actual['flujos'])
                
                    # Mapear periodicidad a texto
                    periodicidad_texto = {
                        1: "anual",
                        2: "semestral", 
                        3: "trimestral",
                        4: "trimestral",
                        6: "bimestral",
                        12: "mensual"
                        }.get(bono_actual['periodicidad'], f"{bono_actual['periodicidad']} veces al año")
                    
                    # Métricas principales — 3 cards agrupadas
                    st.markdown(f'''
                    <div class="metrics-card">
                        <div class="metrics-card-title">Precio y estructura</div>
                        <div class="metrics-row">
                            <div class="metric-card">
                                <div class="metric-label">Precio Limpio</div>
                                <div class="metric-value">{formatear_numero(precio_limpio, 4)}</div>
                            </div>
                            <div class="metric-card">
                                <div class="metric-label">Intereses Corridos</div>
                                <div class="metric-value">{formatear_numero(intereses_corridos, 4)}</div>
                            </div>
                            <div class="metric-card">
                                <div class="metric-label">Capital Residual</div>
                                <div class="metric-value">{formatear_numero(capital_residual, 2)}</div>
                            </div>
                            <div class="metric-card">
                                <div class="metric-label">Cupón Vigente</div>
                                <div class="metric-value">{cupon_vigente:.2%}</div>
                            </div>
                        </div>
                    </div>
                    <div class="metrics-card">
                        <div class="metrics-card-title">Rendimiento y duración</div>
                        <div class="metrics-row">
                            <div class="metric-card">
                                <div class="metric-label">TIR Efectiva</div>
                                <div class="metric-value">{ytm_efectiva:.4%}</div>
                            </div>
                            <div class="metric-card">
                                <div class="metric-label">TIR {periodicidad_texto.title()}</div>
                                <div class="metric-value">{ytm_anualizada:.4%}</div>
                            </div>
                            <div class="metric-card">
                                <div class="metric-label">Duración Modificada</div>
                                <div class="metric-value">{formatear_numero(duracion_modificada, 2)} años</div>
                            </div>
                            <div class="metric-card">
                                <div class="metric-label">Duración Macaulay</div>
                                <div class="metric-value">{formatear_numero(duracion_macaulay, 2)} años</div>
                            </div>
                        </div>
                    </div>
                    <div class="metrics-card">
                        <div class="metrics-card-title">Otros indicadores</div>
                        <div class="metrics-row">
                            <div class="metric-card">
                                <div class="metric-label">Valor Técnico</div>
                                <div class="metric-value">{formatear_numero(valor_tecnico, 4)}</div>
                            </div>
                            <div class="metric-card">
                                <div class="metric-label">Paridad</div>
                                <div class="metric-value">{formatear_numero(paridad, 4)}</div>
                            </div>
                            <div class="metric-card">
                                <div class="metric-label">Próximo Cupón</div>
                                <div class="metric-value">{proximo_cupon.strftime('%d/%m/%Y') if proximo_cupon else 'N/A'}</div>
                            </div>
                            <div class="metric-card">
                                <div class="metric-label">Vida Media</div>
                                <div class="metric-value">{formatear_numero(vida_media, 2)} años</div>
                            </div>
                        </div>
                    </div>
                    ''', unsafe_allow_html=True)
    else:
        # No mostrar nada cuando hay bono seleccionado pero no se ha calculado
        pass
    
    # ── MONITOR ──────────────────────────────────────────────────────────────
    if (st.session_state.get('monitor')
            and not st.session_state.get('bono_seleccionado')
            and not st.session_state.get('flujos_bonos_seleccionados')):

        st.markdown("""<style>
        [data-testid="stSidebar"]{display:none!important}
        [data-testid="stSidebarCollapseButton"]{display:none!important}
        section[data-testid="stMain"]{margin-left:0!important;width:100vw!important;max-width:100vw!important}
        section[data-testid="stMain"] > div:first-child > div:first-child > div:first-child {padding-left:15vw!important;padding-right:15vw!important}
        [data-testid="stMainBlockContainer"]{padding-left:15vw!important;padding-right:15vw!important;max-width:100vw!important}
        </style>""", unsafe_allow_html=True)

        # 6 estados: gráfico + tabla por cada tipo
        _ESTADOS = [
            ('Soberano USD',     'grafico'),
            ('Soberano USD',     'tabla'),
            ('Lecaps & Boncaps', 'grafico'),
            ('Lecaps & Boncaps', 'tabla'),
            ('Bonos CER',        'grafico'),
            ('Bonos CER',        'tabla'),
        ]
        _CYCLE = 30
        _now   = time.time()

        if 'monitor_tick'  not in st.session_state:
            st.session_state.monitor_tick  = _now
            st.session_state.monitor_panel = 0
        if 'monitor_auto'  not in st.session_state:
            st.session_state.monitor_auto  = True

        _auto = st.session_state.monitor_auto

        if _auto and (_now - st.session_state.monitor_tick >= _CYCLE):
            st.session_state.monitor_panel = (st.session_state.monitor_panel + 1) % len(_ESTADOS)
            st.session_state.monitor_tick  = _now

        _remaining = max(1.0, _CYCLE - (_now - st.session_state.monitor_tick))
        _pidx      = st.session_state.monitor_panel
        _tipo, _modo = _ESTADOS[_pidx]

        # Timer de rotación automática — guard en window.parent para evitar acumulación
        _mon_tick_ms = int(_remaining * 1000)
        _ts = int(_now * 1000)
        st.components.v1.html(f"""<script>
        (function() {{
            var ts = {_ts};
            var auto = {'true' if _auto else 'false'};
            // Cancelar timer anterior si existe
            if (window.parent._monTimer) {{
                clearTimeout(window.parent._monTimer);
                window.parent._monTimer = null;
            }}
            if (!auto) return;
            // Evitar registrar el mismo tick dos veces
            if (window.parent._monLastTs === ts) return;
            window.parent._monLastTs = ts;
            window.parent._monTimer = setTimeout(function() {{
                window.parent._monTimer = null;
                var els = window.parent.document.querySelectorAll('button p, button span');
                for (var i = 0; i < els.length; i++) {{
                    if (els[i].textContent.trim() === '\u23f2montick') {{
                        els[i].closest('button').click();
                        return;
                    }}
                }}
            }}, {_mon_tick_ms});
        }})();
        </script>""", height=0)

        # Fetch precios
        with st.spinner("Cargando datos..."):
            _mf = get_next_business_day()
            _mp = {**obtener_precios_data912('arg_bonds'),
                   **obtener_precios_data912('arg_corp'),
                   **obtener_precios_data912('arg_notes')}

        def _mon_tabla(cols, df):
            _hdr = ''.join(
                f'<th style="padding:10px 14px;text-align:{"left" if c == "Activo" else "center"};border-bottom:2px solid #e0e0e0;background:#fafafa;font-weight:700;font-size:16px">{c}</th>'
                for c in cols)
            _rows = ''
            for _i, _r in df[cols].iterrows():
                _bg = '#ffffff' if _i % 2 == 0 else '#f7f9fc'
                _cells = ''
                for c in cols:
                    val = _r[c]
                    style = f'padding:9px 14px;border-bottom:1px solid #e8e8e8;font-size:16px;text-align:{"left" if c == "Activo" else "center"}'
                    if c == 'Var. %' and isinstance(val, str) and val != '-':
                        try:
                            num = float(val.replace('%','').replace('+',''))
                            style += ';color:' + ('#2e7d32' if num > 0 else '#c62828' if num < 0 else '#333')
                            style += ';font-weight:600'
                        except: pass
                    _cells += f'<td style="{style}">{val}</td>'
                _rows += f'<tr style="background:{_bg}">{_cells}</tr>'
            st.markdown(f'<div style="border-radius:8px;overflow:hidden;border:1px solid #e0e0e0;margin-top:8px"><table style="width:100%;border-collapse:collapse">{_hdr}{_rows}</table></div>', unsafe_allow_html=True)

        st.markdown(f"<h2 style='text-align:center;margin:8px 0 16px'>{_tipo}</h2>", unsafe_allow_html=True)

        # ── Soberano USD ──────────────────────────────────────────────────
        if _tipo == 'Soberano USD':
            _filas = []
            def _mon_grupo_sov(ticker):
                t = ticker.upper()
                if t.startswith('GD'): return 'GD'
                elif t.startswith('B'): return 'B'
                else: return 'AL'
            for _b in bonos:
                if _b.get('tipo_bono') != 'Soberano USD': continue
                _tk = _b.get('ticker', '').strip().upper()
                if not _tk: continue
                _tkapi = (_tk + 'D') if len(_tk) == 4 else _tk
                _pd = _mp.get(_tkapi)
                if not _pd or not _pd.get('c') or _pd['c'] <= 0: continue
                _precio = _pd['c']; _pct = _pd.get('pct_change')
                try:
                    _fl = [f['total'] for f in _b['flujos'] if f['fecha'] > _mf]
                    _fe = [f['fecha']  for f in _b['flujos'] if f['fecha'] > _mf]
                    if not _fl: continue
                    _ytm  = calcular_ytm(_precio, _fl, _fe, _mf, _b['base_calculo'])
                    _tirs = _b['periodicidad'] * ((1 + _ytm) ** (1 / _b['periodicidad']) - 1)
                    _mac  = calcular_duracion_macaulay(_fl, _fe, _mf, _ytm, _b['base_calculo'])
                    _ytma = _b['periodicidad'] * ((1 + _ytm) ** (1 / _b['periodicidad']) - 1)
                    _dur  = calcular_duracion_modificada(_mac, _ytma / _b['periodicidad'], _b['periodicidad'])
                    _vcto = encontrar_fecha_vencimiento(_b['flujos'])
                    _filas.append({
                        'Activo': _b['nombre'], 'Ticker': _tk,
                        'Vencimiento': _vcto.strftime('%d/%m/%Y') if _vcto else '-',
                        'Precio': f"{_precio:.2f}",
                        'TIR Semestral': f"{_tirs*100:.2f}%",
                        'Dur. Mod.': f"{_dur:.2f}",
                        'Var. %': f"{_pct:+.2f}%" if _pct is not None and not pd.isna(_pct) else '-',
                        '_tir': _tirs * 100, '_dur': _dur,
                        '_grupo': _mon_grupo_sov(_tk),
                    })
                except: continue
            if _filas:
                _df = pd.DataFrame(_filas)
                if _modo == 'grafico':
                    _dc = _df[['Activo', 'Ticker', '_grupo', '_dur', '_tir']].dropna()
                    _dc = _dc[_dc['_dur'] > 0]
                    _fig = go.Figure()
                    for _gs, _cp, _cl, _nm in [
                        ('GD', '#1a237e', '#42a5f5', 'Globales'),
                        ('AL', '#1565c0', '#90caf9', 'Bonares'),
                    ]:
                        _ds = _dc[_dc['_grupo'] == _gs]
                        if len(_ds) < 2: continue
                        _xg, _yg = _ds['_dur'].values, _ds['_tir'].values
                        _coeffs = np.polyfit(np.log(_xg), _yg, 1)
                        _xl = np.linspace(_xg.min(), _xg.max(), 200)
                        _fig.add_trace(go.Scatter(
                            x=_xg, y=_yg, mode='markers+text',
                            text=_ds['Activo'], textposition='top center',
                            textfont=dict(size=14, color=_cp),
                            marker=dict(size=10, color=_cp),
                            name=_nm,
                            hovertemplate='<b>%{text}</b><br>Dur. Mod.: %{x:.2f}<br>TIR Sem.: %{y:.2f}%<extra></extra>'))
                        _fig.add_trace(go.Scatter(
                            x=_xl, y=_coeffs[0]*np.log(_xl)+_coeffs[1], mode='lines',
                            line=dict(color=_cl, width=2, dash='dash'),
                            hoverinfo='skip', showlegend=False))
                    _db = _dc[_dc['_grupo'] == 'B']
                    if len(_db) > 0:
                        _fig.add_trace(go.Scatter(
                            x=_db['_dur'].values, y=_db['_tir'].values,
                            mode='markers+text', text=_db['Activo'], textposition='top center',
                            textfont=dict(size=14, color='#6a1b9a'),
                            marker=dict(size=10, color='#6a1b9a'),
                            name='Bopreal',
                            hovertemplate='<b>%{text}</b><br>Dur. Mod.: %{x:.2f}<br>TIR Sem.: %{y:.2f}%<extra></extra>'))
                    _fig.update_layout(
                        title='Cueva de Rendimientos — Soberano USD',
                        xaxis_title='Duración Modificada', yaxis_title='TIR Semestral (%)',
                        plot_bgcolor='#f4f6fb', paper_bgcolor='#f4f6fb', height=500,
                        margin=dict(t=60, b=40, l=60, r=20),
                        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
                        xaxis=dict(showgrid=True, gridcolor='#cccccc'),
                        yaxis=dict(showgrid=True, gridcolor='#cccccc'))
                    st.plotly_chart(_fig, use_container_width=True)
                else:
                    _mon_tabla(['Activo','Vencimiento','Precio','TIR Semestral','Dur. Mod.','Var. %'], _df)
            else:
                st.info("No hay precios disponibles para Soberano USD.")

        # ── Lecaps & Boncaps ──────────────────────────────────────────────
        elif _tipo == 'Lecaps & Boncaps':
            _lf    = _mf.date() if hasattr(_mf, 'date') else _mf
            _filas = []
            for _b in bonos:
                if _b.get('tipo_bono') != 'Lecaps & Boncaps': continue
                _tk = _b.get('ticker', '').strip().upper()
                if not _tk: continue
                _pd = _mp.get(_tk)
                if not _pd or not _pd.get('c') or _pd['c'] <= 0: continue
                _precio = _pd['c']; _pct = _pd.get('pct_change')
                _mat  = _b.get('maturity')
                if not _mat: continue
                _matd = _mat.date() if hasattr(_mat, 'date') else _mat
                _dr   = max((_matd - _lf).days, 0)
                if _dr == 0: continue
                _vf  = _b.get('valor_final', 0) or 0
                _tna = (_vf - _precio) / _precio / _dr * 365 if _precio > 0 else None
                _tea = (1 + (_vf - _precio) / _precio) ** (365.0 / _dr) - 1 if _precio > 0 else None
                _tem = (1 + _tea) ** (1/12) - 1 if _tea is not None else None
                _filas.append({
                    'Activo': _b['nombre'],
                    'Vencimiento': _matd.strftime('%d/%m/%Y'),
                    'Días Rem.': _dr,
                    'Precio': f"{_precio:.2f}",
                    'TNA': f"{_tna*100:.2f}%" if _tna is not None else '-',
                    'TEM': f"{_tem*100:.2f}%" if _tem is not None else '-',
                    'Var. %': f"{_pct:+.2f}%" if _pct is not None and not pd.isna(_pct) else '-',
                    '_tna': _tna * 100 if _tna else None, '_dr': float(_dr),
                })
            if _filas:
                _df = pd.DataFrame(_filas).sort_values('Días Rem.').reset_index(drop=True)
                _lc = _df[['Activo', '_dr', '_tna']].dropna()
                _lc = _lc[_lc['_tna'] > 0]
                if _modo == 'grafico':
                    if len(_lc) >= 2:
                        _x_lec, _y_lec = _lc['_dr'].values, _lc['_tna'].values
                        _fig = go.Figure()
                        _fig.add_trace(go.Scatter(
                            x=_x_lec, y=_y_lec,
                            mode='markers+text', text=_lc['Activo'], textposition='top center',
                            textfont=dict(size=14, color='#1a237e'),
                            marker=dict(size=10, color='#1a237e'),
                            hovertemplate='<b>%{text}</b><br>Días: %{x:.0f}<br>TNA: %{y:.2f}%<extra></extra>'))
                        if len(_lc) >= 3:
                            _cl = np.polyfit(np.log(_x_lec), _y_lec, 1)
                            _xll = np.linspace(_x_lec.min(), _x_lec.max(), 100)
                            _fig.add_trace(go.Scatter(
                                x=_xll, y=_cl[0]*np.log(_xll)+_cl[1], mode='lines',
                                line=dict(color='#42a5f5', width=2, dash='dash'),
                                hoverinfo='skip', showlegend=False))
                        _fig.update_layout(
                            title='Cueva de Rendimientos — Lecaps & Boncaps',
                            xaxis_title='Días al Vencimiento', yaxis_title='TNA (%)',
                            plot_bgcolor='#f4f6fb', paper_bgcolor='#f4f6fb', height=500,
                            margin=dict(t=50, b=40, l=60, r=20),
                            xaxis=dict(showgrid=True, gridcolor='#cccccc'),
                            yaxis=dict(showgrid=True, gridcolor='#cccccc'))
                        st.plotly_chart(_fig, use_container_width=True)
                    else:
                        st.info("No hay suficientes datos para el gráfico.")
                else:
                    _mon_tabla(['Activo','Vencimiento','Días Rem.','Precio','TNA','TEM','Var. %'], _df)
            else:
                st.info("No hay precios disponibles para Lecaps & Boncaps.")

        # ── Bonos CER ─────────────────────────────────────────────────────
        else:
            _lf = _mf.date() if hasattr(_mf, 'date') else _mf
            _cer_mon, _ = obtener_cer_settlement(_mf)
            _filas = []
            for _b in bonos:
                if _b.get('tipo_bono') != 'Bonos CER': continue
                _tk = _b.get('ticker', '').strip().upper()
                if not _tk: continue
                _pd = _mp.get(_tk)
                if not _pd or not _pd.get('c') or _pd['c'] <= 0: continue
                _precio = _pd['c']; _pct = _pd.get('pct_change')
                _mat = _b.get('maturity')
                if not _mat: continue
                _matd = _mat.date() if hasattr(_mat, 'date') else _mat
                _dr = max((_matd - _lf).days, 0)
                if _dr == 0: continue
                _cb = _b.get('cer_base') or 0
                _fcv = round(_cer_mon / _cb, 4) if (_cer_mon and _cb) else None
                if _fcv and _precio > 0 and _dr > 0:
                    _tir_a = (_fcv * 100 / _precio) ** (365.0 / _dr) - 1
                    _tir_m = (1 + _tir_a) ** (30 / 360) - 1
                    _vm = _dr / 365.0
                    _dur_m = _vm / (1 + _tir_a)
                else:
                    _tir_a = _tir_m = _dur_m = None
                _filas.append({
                    'Activo': _b['nombre'],
                    'Vencimiento': _matd.strftime('%d/%m/%Y'),
                    'Días Rem.': _dr,
                    'Precio': f"{_precio:.2f}",
                    'TIR Anual': f"{_tir_a*100:.2f}%" if _tir_a is not None else '-',
                    'TIR Mensual': f"{_tir_m*100:.2f}%" if _tir_m is not None else '-',
                    'Dur. Mod.': f"{_dur_m:.2f}" if _dur_m is not None else '-',
                    'Var. %': f"{_pct:+.2f}%" if _pct is not None and not pd.isna(_pct) else '-',
                    '_tir_a': _tir_a * 100 if _tir_a is not None else None,
                    '_dur_m': _dur_m,
                })
            if _filas:
                _df = pd.DataFrame(_filas).sort_values('Días Rem.').reset_index(drop=True)
                _cc = _df[['Activo', '_dur_m', '_tir_a']].dropna()
                _cc = _cc[_cc['_dur_m'] > 0]
                if _modo == 'grafico':
                    if len(_cc) >= 2:
                        _x_cer, _y_cer = _cc['_dur_m'].values, _cc['_tir_a'].values
                        _fig = go.Figure()
                        _fig.add_trace(go.Scatter(
                            x=_x_cer, y=_y_cer,
                            mode='markers+text', text=_cc['Activo'], textposition='top center',
                            textfont=dict(size=14, color='#1a237e'),
                            marker=dict(size=10, color='#1a237e'),
                            hovertemplate='<b>%{text}</b><br>Dur. Mod.: %{x:.2f}<br>TIR Anual: %{y:.2f}%<extra></extra>'))
                        if len(_cc) >= 3:
                            _cc2 = np.polyfit(np.log(_x_cer), _y_cer, 1)
                            _xlc = np.linspace(_x_cer.min(), _x_cer.max(), 100)
                            _fig.add_trace(go.Scatter(
                                x=_xlc, y=_cc2[0]*np.log(_xlc)+_cc2[1], mode='lines',
                                line=dict(color='#42a5f5', width=2, dash='dash'),
                                hoverinfo='skip', showlegend=False))
                        _fig.update_layout(
                            title='Cueva de Rendimientos — Bonos CER',
                            xaxis_title='Duración Modificada', yaxis_title='TIR Anual (%)',
                            plot_bgcolor='#f4f6fb', paper_bgcolor='#f4f6fb', height=500,
                            margin=dict(t=50, b=40, l=60, r=20),
                            xaxis=dict(showgrid=True, gridcolor='#cccccc'),
                            yaxis=dict(showgrid=True, gridcolor='#cccccc'))
                        st.plotly_chart(_fig, use_container_width=True)
                    else:
                        st.info("No hay suficientes datos para el gráfico.")
                else:
                    _mon_tabla(['Activo','Vencimiento','Días Rem.','Precio','TIR Anual','TIR Mensual','Dur. Mod.','Var. %'], _df)
            else:
                st.info("No hay precios disponibles para Bonos CER.")

        # Botones al final — margen grande para quedar fuera de la vista
        st.markdown("<div style='margin-top:120px'></div>", unsafe_allow_html=True)
        _bca, _bcb, _bcc, _bcd, _bce = st.columns([1, 1, 1, 1, 2])
        with _bca:
            if st.button("← Anterior", key="mon_prev", use_container_width=True):
                st.session_state.monitor_panel = (st.session_state.monitor_panel - 1) % len(_ESTADOS)
                st.session_state.monitor_tick  = time.time()
                st.rerun()
        with _bcb:
            if st.button("Siguiente →", key="mon_next", use_container_width=True):
                st.session_state.monitor_panel = (st.session_state.monitor_panel + 1) % len(_ESTADOS)
                st.session_state.monitor_tick  = time.time()
                st.rerun()
        with _bcc:
            _lbl_auto = "⏸ Pausar" if _auto else "▶ Reanudar"
            if st.button(_lbl_auto, key="mon_auto_toggle", use_container_width=True):
                st.session_state.monitor_auto = not _auto
                st.session_state.monitor_tick = time.time()
                st.rerun()
        with _bcd:
            if st.button("Salir del Monitor", type="secondary", use_container_width=True, key="mon_exit"):
                st.session_state.monitor = False
                st.session_state.bono_seleccionado        = st.session_state.pop('monitor_prev_bono', None)
                st.session_state.flujos_bonos_seleccionados = st.session_state.pop('monitor_prev_flujos', [])
                for _k in ('monitor_tick', 'monitor_panel', 'monitor_auto'):
                    st.session_state.pop(_k, None)
                st.rerun()
        # Botón hidden que el timer JS clickea para disparar rerun
        with _bce:
            if st.button("\u23f2montick", key="mon_tick_hidden"):
                pass
        # Ocultar el botón tick via JS (CSS no puede seleccionar por texto)
        st.components.v1.html("""<script>
        (function hide() {
            var els = window.parent.document.querySelectorAll('button p, button span');
            for (var i = 0; i < els.length; i++) {
                if (els[i].textContent.trim() === '\u23f2montick') {
                    els[i].closest('button').style.display = 'none';
                    return;
                }
            }
            setTimeout(hide, 100);
        })();
        </script>""", height=0)

    # ── S0: Tablas de mercado ─────────────────────────────────────────────
    if not st.session_state.get('bono_seleccionado') and not st.session_state.get('flujos_bonos_seleccionados') and not st.session_state.get('monitor'):
        # Auto-refresh cada 10 minutos (con guard para evitar timers acumulados)
        st.components.v1.html("""<script>
        if (!window._autoRefreshSet) {
            window._autoRefreshSet = true;
            setTimeout(function(){ window.location.reload(); }, 600000);
        }
        </script>""", height=0)

        TABLE_CSS = """
        <style>
        .bond-wrap { border-radius:10px; overflow:hidden; border:1px solid #e0e0e0; }
        .bond-title { background:#fafafa; color:#333; font-weight:700; font-size:14px; padding:11px 14px; border-bottom:2px solid #e0e0e0; font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif; letter-spacing:0.02em; }
        .bond-table { width:100%; border-collapse:collapse; font-size:13px; font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif; }
        .bond-table th { background:#fafafa; color:#555; font-weight:600; padding:9px 12px; text-align:center; border-bottom:2px solid #e0e0e0; white-space:nowrap; }
        .bond-table th:first-child { text-align:left; }
        .bond-table td { padding:8px 12px; color:#333; white-space:nowrap; text-align:center; }
        .bond-table td:first-child { text-align:left; }
        .bond-table tr:nth-child(even) td { background:#f7f7f7; }
        .bond-table tr:nth-child(odd) td { background:#ffffff; }
        .bond-table tr:hover td { background:#eef2ff; }
        </style>
        """

        def _esc(v):
            return str(v).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;')

        def render_tabla_html(df, titulo='', separadores=None):
            # separadores: set of row indices before which to insert a visible separator row
            sep_set = set(separadores) if separadores else set()
            cols = list(df.columns)
            ncols = len(cols)
            headers = ''.join(
                '<th style="width:40px"></th>' if c == '_manual_control' else f'<th>{_esc(c)}</th>'
                for c in cols
            )
            sep_row = f'<tr><td colspan="{ncols}" style="height:2px;background:#b0bec5;padding:0;border:none;line-height:0;font-size:0;"></td></tr>'
            rows = ''
            for idx, (_, row) in enumerate(df.iterrows()):
                if idx in sep_set:
                    rows += sep_row
                cells = ''
                for col in cols:
                    val = row[col]
                    if col == '_manual_control':
                        cells += f'<td style="text-align:center">{val}</td>'
                        continue
                    val_str = _esc(val)
                    if col == 'Var. Diaria %' and val != '-':
                        color = '#2e7d32' if str(val).startswith('+') else '#c62828'
                        cells += f'<td style="color:{color};font-weight:600">{val_str}</td>'
                    else:
                        cells += f'<td>{val_str}</td>'
                rows += f'<tr>{cells}</tr>'
            title_html = f'<div class="bond-title">{_esc(titulo)}</div>' if titulo else ''
            return f'<div class="bond-wrap">{title_html}<table class="bond-table"><thead><tr>{headers}</tr></thead><tbody>{rows}</tbody></table></div>'

        def _guardar_precio_manual_monitor(tabla_id, row_id, precio):
            state_key = _manual_price_state_key(tabla_id, row_id)
            if precio is None:
                st.session_state.pop(state_key, None)
            else:
                st.session_state[state_key] = precio

        def _obtener_precio_manual_tabla(tabla_id, row_id):
            return normalizar_precio_manual_monitor(st.session_state.get(_manual_price_state_key(tabla_id, row_id)))

        def _close_manual_dialog():
            st.session_state.pop("monitor_manual_dialog_table", None)
            st.session_state.pop("monitor_manual_dialog_activo", None)

        @st.dialog(" ", width="small", dismissible=True, on_dismiss=_close_manual_dialog)
        def _manual_price_dialog():
            tabla_id = st.session_state.get("monitor_manual_dialog_table")
            activo = st.session_state.get("monitor_manual_dialog_activo")
            precio_mercado = float(st.session_state.get("monitor_manual_dialog_precio_mercado", 0.0) or 0.0)
            precio_manual = _obtener_precio_manual_tabla(tabla_id, activo) if tabla_id and activo else None
            if not tabla_id or not activo:
                return
            input_key = "monitor_manual_dialog_input"
            activo_key = "monitor_manual_dialog_input_activo"
            if st.session_state.get(activo_key) != f"{tabla_id}:{activo}":
                st.session_state[input_key] = float(precio_manual if precio_manual is not None else precio_mercado)
                st.session_state[activo_key] = f"{tabla_id}:{activo}"

            def _apply_manual_dialog():
                _guardar_precio_manual_monitor(
                    tabla_id,
                    activo,
                    normalizar_precio_manual_monitor(st.session_state.get(input_key))
                )
                _close_manual_dialog()

            st.number_input(
                "",
                min_value=0.0,
                step=0.01,
                format="%.2f",
                key=input_key,
                label_visibility="collapsed",
                on_change=_apply_manual_dialog,
            )

        qp_toggle = st.query_params.get("manual_toggle")
        qp_table = st.query_params.get("manual_table")
        qp_market = st.query_params.get("manual_market")
        if qp_toggle and qp_table:
            tabla_id_qp = qp_table if isinstance(qp_table, str) else qp_table[-1]
            activo_qp = qp_toggle if isinstance(qp_toggle, str) else qp_toggle[-1]
            precio_market_qp = qp_market if isinstance(qp_market, str) else (qp_market[-1] if qp_market else "0")
            if _obtener_precio_manual_tabla(tabla_id_qp, activo_qp) is not None:
                _guardar_precio_manual_monitor(tabla_id_qp, activo_qp, None)
                _close_manual_dialog()
            else:
                st.session_state["monitor_manual_dialog_table"] = tabla_id_qp
                st.session_state["monitor_manual_dialog_activo"] = activo_qp
                try:
                    st.session_state["monitor_manual_dialog_precio_mercado"] = float(precio_market_qp)
                except (TypeError, ValueError):
                    st.session_state["monitor_manual_dialog_precio_mercado"] = 0.0
            st.query_params.clear()

        if st.session_state.get("monitor_manual_dialog_table") and st.session_state.get("monitor_manual_dialog_activo"):
            _manual_price_dialog()

        # --- Fetch precios y calcular grupos (compartido entre tabs) ---
        with st.spinner("Cargando precios y calculando métricas..."):
            fecha_hoy = get_next_business_day()
            precios_bonds = obtener_precios_data912('arg_bonds')
            precios_corp = obtener_precios_data912('arg_corp')
            precios_todos = {**precios_bonds, **precios_corp}

            grupos = {}
            for bono in bonos:
                ticker = bono.get('ticker', '').strip()
                if not ticker or ticker == 'SPX500':
                    continue

                ticker_api = ticker.upper()
                if len(ticker_api) == 4:
                    ticker_api = ticker_api + 'D'

                precio_data = precios_todos.get(ticker_api)
                if not precio_data:
                    continue
                precio = precio_data['c']
                pct_change = precio_data.get('pct_change')
                if not precio or precio <= 0:
                    continue

                try:
                    precio_manual = obtener_precio_manual_monitor(bono.get('tipo_bono', 'Otros'), bono['nombre'])
                    precio_calculo = precio_manual if precio_manual is not None else precio
                    flujos = []
                    fechas = []
                    for flujo in bono['flujos']:
                        if flujo['fecha'] > fecha_hoy:
                            flujos.append(flujo['total'])
                            fechas.append(flujo['fecha'])

                    if not flujos:
                        continue

                    capital_residual = 100 - sum([
                        f['capital'] for f in bono['flujos'] if f['fecha'] <= fecha_hoy
                    ])

                    todas_fechas = [f['fecha'] for f in bono['flujos']]
                    fecha_ultimo_cupon = encontrar_ultimo_cupon(fecha_hoy, todas_fechas, bono.get('fecha_emision'))
                    intereses_corridos = 0
                    if fecha_ultimo_cupon:
                        intereses_corridos = calcular_intereses_corridos(
                            fecha_hoy, fecha_ultimo_cupon,
                            bono['tasa_cupon'], capital_residual, bono['base_calculo']
                        )

                    ytm_efectiva = calcular_ytm(
                        precio_calculo, flujos, fechas, fecha_hoy,
                        bono['base_calculo'], bono['periodicidad']
                    )
                    if (1 + ytm_efectiva) <= 0:
                        continue
                    tir_semestral = 2 * ((1 + ytm_efectiva) ** (1/2) - 1)

                    duracion_macaulay = calcular_duracion_macaulay(
                        flujos, fechas, fecha_hoy, ytm_efectiva, bono['base_calculo']
                    )
                    ytm_anualizada = bono['periodicidad'] * ((1 + ytm_efectiva) ** (1 / bono['periodicidad']) - 1)
                    duracion_modificada = calcular_duracion_modificada(
                        duracion_macaulay,
                        ytm_anualizada / bono['periodicidad'],
                        bono['periodicidad']
                    )

                    cupon_vigente = encontrar_cupon_vigente(fecha_hoy, bono['flujos'])

                    tipo = bono.get('tipo_bono', 'Otros')
                    fecha_vcto = encontrar_fecha_vencimiento(bono['flujos'])
                    fila = {
                        'Activo': bono['nombre'],
                        'Ticker': ticker,
                        'Vencimiento': fecha_vcto.strftime('%d/%m/%Y') if fecha_vcto else '-',
                        'Precio': precio_calculo,
                        '_precio_mercado': precio,
                        'Int. Corridos': round(intereses_corridos, 4),
                        'Cap. Residual': round(capital_residual, 2),
                        'Cupón Vigente': round(cupon_vigente * 100, 4),
                        'TIR Semestral': round(tir_semestral * 100, 2),
                        'Dur. Modificada': round(duracion_modificada, 2),
                        'Var. Diaria %': pct_change,
                    }
                    if tipo not in grupos:
                        grupos[tipo] = []
                    grupos[tipo].append(fila)

                except Exception:
                    continue

        def _render_grupos(tipos_a_mostrar):
            tipos_presentes = [t for t in tipos_a_mostrar if grupos.get(t)]
            if not tipos_presentes:
                st.info("No hay precios disponibles en este momento.")
                return
            st.markdown(TABLE_CSS, unsafe_allow_html=True)
            for tipo in tipos_presentes:
                st.markdown("<div style='margin-top:2rem'></div>", unsafe_allow_html=True)
                st.markdown(f'<div class="bond-wrap"><div class="bond-title">{tipo}</div></div>', unsafe_allow_html=True)
                df_tabla = pd.DataFrame(grupos[tipo])
                if 'corporativo' in tipo.lower():
                    df_tabla = df_tabla.sort_values(['Activo', 'Dur. Modificada']).reset_index(drop=True)
                df_raw = df_tabla.copy()
                _sov_separadores = set()

                # Cueva Soberano USD
                if 'soberano' in tipo.lower():
                    def _sov_grupo(ticker):
                        t = ticker.upper()
                        if t.startswith('GD'): return 'GD'
                        elif t.startswith('B'): return 'B'
                        else: return 'AL'
                    df_tabla['_grupo'] = df_tabla['Ticker'].apply(_sov_grupo)
                    orden_grupo = {'AL': 0, 'GD': 1, 'B': 2}
                    df_tabla['_orden_grupo'] = df_tabla['_grupo'].map(orden_grupo)
                    df_tabla = df_tabla.sort_values(['_orden_grupo', 'Dur. Modificada']).reset_index(drop=True)
                    df_raw = df_tabla.copy()
                    # Calcular índices de separación entre grupos para la tabla
                    _grupos_col = df_tabla['_grupo'].tolist()
                    _sov_separadores = {0} | {i for i in range(1, len(_grupos_col)) if _grupos_col[i] != _grupos_col[i-1]}
                    df_tabla = df_tabla.drop(columns=['_grupo', '_orden_grupo'])
                    df_curva = df_raw[['Activo', 'Ticker', '_grupo', 'Dur. Modificada', 'TIR Semestral']].dropna()
                    df_curva = df_curva[df_curva['Dur. Modificada'] > 0]
                    df_gd = df_curva[df_curva['_grupo'] == 'GD']
                    df_al = df_curva[df_curva['_grupo'] == 'AL']
                    df_b  = df_curva[df_curva['_grupo'] == 'B']
                    if len(df_curva) >= 3:
                        fig = go.Figure()
                        for df_serie, color_pt, color_ln, nombre in [
                            (df_gd, '#1a237e', '#42a5f5', 'GD (Ley NY)'),
                            (df_al, '#1565c0', '#90caf9', 'AL / AO / AN (Ley ARG)'),
                        ]:
                            if len(df_serie) < 2:
                                continue
                            x = df_serie['Dur. Modificada'].values
                            y = df_serie['TIR Semestral'].values
                            coeffs = np.polyfit(np.log(x), y, 1)
                            x_line = np.linspace(x.min(), x.max(), 200)
                            y_line = coeffs[0] * np.log(x_line) + coeffs[1]
                            fig.add_trace(go.Scatter(x=x, y=y, mode='markers+text',
                                text=df_serie['Activo'], textposition='top center',
                                textfont=dict(size=10, color=color_pt), marker=dict(size=9, color=color_pt),
                                name=nombre, hovertemplate='<b>%{text}</b><br>Dur. Mod.: %{x:.2f}<br>TIR Sem.: %{y:.2f}%<extra></extra>'))
                            fig.add_trace(go.Scatter(x=x_line, y=y_line, mode='lines',
                                line=dict(color=color_ln, width=2, dash='dash'),
                                name=f'Tend. {nombre}', hoverinfo='skip', showlegend=False))
                        if len(df_b) > 0:
                            fig.add_trace(go.Scatter(
                                x=df_b['Dur. Modificada'].values, y=df_b['TIR Semestral'].values,
                                mode='markers+text', text=df_b['Activo'], textposition='top center',
                                textfont=dict(size=10, color='#1a237e'),
                                marker=dict(size=9, color='#1a237e', symbol='circle'),
                                name='Bonares (B)',
                                hovertemplate='<b>%{text}</b><br>Dur. Mod.: %{x:.2f}<br>TIR Sem.: %{y:.2f}%<extra></extra>'))
                        fig.update_layout(
                            title='Cueva de Rendimientos — Soberano USD',
                            xaxis_title='Duración Modificada (años)', yaxis_title='TIR Semestral (%)',
                            xaxis=dict(showgrid=True, gridcolor='#cccccc', linecolor='#999999', linewidth=1, showline=True, tickfont=dict(color='#444444'), title_font=dict(color='#444444')),
                            yaxis=dict(showgrid=True, gridcolor='#cccccc', linecolor='#999999', linewidth=1, showline=True, tickfont=dict(color='#444444'), title_font=dict(color='#444444')),
                            plot_bgcolor='#f4f6fb', paper_bgcolor='#f4f6fb',
                            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
                            height=420, margin=dict(t=60, b=40, l=60, r=20))
                        st.plotly_chart(fig, use_container_width=True)

                # Cueva Corporativos
                if 'corporativo' in tipo.lower():
                    titulo_cueva = f'Cueva de Rendimientos — {tipo}'
                    df_curva = df_raw[['Activo', 'Dur. Modificada', 'TIR Semestral']].dropna()
                    df_curva = df_curva[df_curva['Dur. Modificada'] > 0]
                    df_curva = df_curva[(df_curva['TIR Semestral'] >= 3) & (df_curva['TIR Semestral'] <= 15)]
                    if len(df_curva) >= 3:
                        x = df_curva['Dur. Modificada'].values
                        y = df_curva['TIR Semestral'].values
                        coeffs = np.polyfit(np.log(x), y, 1)
                        x_line = np.linspace(x.min(), x.max(), 200)
                        y_line = coeffs[0] * np.log(x_line) + coeffs[1]
                        x_r = x.max() - x.min() if x.max() != x.min() else 1
                        y_r = y.max() - y.min() if y.max() != y.min() else 1
                        nombres = df_curva['Activo'].values
                        y_trend_pts = coeffs[0] * np.log(x) + coeffs[1]
                        desviacion = np.abs(y - y_trend_pts)
                        orden = np.argsort(desviacion)[::-1]
                        MIN_DIST = 0.11
                        dirs = ['top center','bottom center','middle right','middle left',
                                'top right','top left','bottom right','bottom left']
                        offsets = {'top center':(0,.06),'bottom center':(0,-.06),
                                   'middle right':(.07,0),'middle left':(-.07,0),
                                   'top right':(.05,.05),'top left':(-.05,.05),
                                   'bottom right':(.05,-.05),'bottom left':(-.05,-.05)}
                        texto_visible = [''] * len(x)
                        positions = ['top center'] * len(x)
                        placed = []
                        for i in orden:
                            nx, ny = x[i]/x_r, y[i]/y_r
                            dists_to_placed = [np.sqrt((nx-px)**2+(ny-py)**2) for px,py in placed]
                            if dists_to_placed and min(dists_to_placed) < MIN_DIST:
                                continue
                            best_pos, best_score = 'top center', -1
                            best_lx, best_ly = nx, ny + 0.06
                            for d in dirs:
                                ox, oy = offsets[d]
                                lx, ly = nx + ox, ny + oy
                                dists = [np.sqrt((lx-px)**2+(ly-py)**2) for px,py in placed]
                                dists += [np.sqrt(((x[i]-x[j])/x_r)**2+((y[i]-y[j])/y_r)**2)
                                          for j in range(len(x)) if j != i]
                                score = min(dists) if dists else 1
                                if score > best_score:
                                    best_score, best_pos = score, d
                                    best_lx, best_ly = lx, ly
                            texto_visible[i] = nombres[i]
                            positions[i] = best_pos
                            placed.append((best_lx, best_ly))
                        fig_corp = go.Figure()
                        fig_corp.add_trace(go.Scatter(x=x, y=y, mode='markers+text',
                            text=texto_visible, customdata=nombres, textposition=positions,
                            textfont=dict(size=9, color='#1a237e'), marker=dict(size=9, color='#1a237e'),
                            name=tipo, hovertemplate='<b>%{customdata}</b><br>Dur. Mod.: %{x:.2f}<br>TIR Sem.: %{y:.2f}%<extra></extra>'))
                        fig_corp.add_trace(go.Scatter(x=x_line, y=y_line, mode='lines',
                            line=dict(color='#42a5f5', width=2, dash='dash'),
                            name='Tendencia log.', hoverinfo='skip', showlegend=False))
                        fig_corp.update_layout(
                            title=titulo_cueva,
                            xaxis_title='Duración Modificada (años)', yaxis_title='TIR Semestral (%)',
                            xaxis=dict(showgrid=True, gridcolor='#cccccc', linecolor='#999999', linewidth=1, showline=True, tickfont=dict(color='#444444'), title_font=dict(color='#444444')),
                            yaxis=dict(showgrid=True, gridcolor='#cccccc', linecolor='#999999', linewidth=1, showline=True, tickfont=dict(color='#444444'), title_font=dict(color='#444444')),
                            plot_bgcolor='#f4f6fb', paper_bgcolor='#f4f6fb',
                            legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
                            height=560, margin=dict(t=60, b=40, l=60, r=40))
                        st.plotly_chart(fig_corp, use_container_width=True)

                # Tabla
                row_ids = df_tabla['Activo'].tolist()
                df_tabla_display = df_tabla.drop(columns=['_precio_mercado'], errors='ignore').copy()
                tabla_id = _tabla_manual_id(tipo)
                df_tabla_display['_manual_control'] = df_tabla['Activo'].apply(
                    lambda activo: (
                        f'<a href="?manual_table={tabla_id}&manual_toggle={activo}&manual_market='
                        f'{float(df_tabla.loc[df_tabla["Activo"] == activo, "_precio_mercado"].iloc[0]):.6f}" '
                        'style="text-decoration:none;font-size:16px;color:#2e7d32">●</a>'
                        if _obtener_precio_manual_tabla(tabla_id, activo) is not None else
                        f'<a href="?manual_table={tabla_id}&manual_toggle={activo}&manual_market='
                        f'{float(df_tabla.loc[df_tabla["Activo"] == activo, "_precio_mercado"].iloc[0]):.6f}" '
                        'style="text-decoration:none;font-size:16px;color:#9aa5b1">○</a>'
                    )
                )
                df_tabla_display['Precio'] = df_tabla_display['Precio'].map('{:.2f}'.format)
                df_tabla_display['Int. Corridos'] = df_tabla_display['Int. Corridos'].map('{:.4f}'.format)
                df_tabla_display['Cap. Residual'] = df_tabla_display['Cap. Residual'].map('{:.2f}'.format)
                df_tabla_display['Cupón Vigente'] = df_tabla_display['Cupón Vigente'].map('{:.4f}%'.format)
                df_tabla_display['TIR Semestral'] = df_tabla_display['TIR Semestral'].map('{:.2f}%'.format)
                df_tabla_display['Dur. Modificada'] = df_tabla_display['Dur. Modificada'].map('{:.2f}'.format)
                df_tabla_display['Var. Diaria %'] = df_tabla_display['Var. Diaria %'].apply(
                    lambda x: f'{x:+.2f}%' if x is not None and not pd.isna(x) else '-'
                )
                _sep = _sov_separadores if 'soberano' in tipo.lower() else None
                st.markdown(render_tabla_html(df_tabla_display, separadores=_sep), unsafe_allow_html=True)

        tab_usd, tab_ars, tab_corp = st.tabs(["Soberano - USD", "Soberano - ARS", "Corporativos - USD"])

        with tab_usd:
            _render_grupos(['Soberano USD'])

        with tab_corp:
            _render_grupos(['Corporativo Ley NY', 'Corporativo Ley ARG'])

        with tab_ars:
            with st.spinner("Cargando precios y calculando métricas..."):
                _fhp = get_next_business_day()
                fecha_hoy_p = _fhp.date() if hasattr(_fhp, 'date') else _fhp
                precios_bonds_p = obtener_precios_data912('arg_bonds')
                precios_corp_p  = obtener_precios_data912('arg_corp')
                precios_notes_p = obtener_precios_data912('arg_notes')
                precios_todos_p = {**precios_bonds_p, **precios_corp_p, **precios_notes_p}

                lecaps = [b for b in bonos if b.get('tipo_bono') == 'Lecaps & Boncaps']
                filas_lecap = []
                for bono in lecaps:
                    ticker = bono.get('ticker', '').strip().upper()
                    if not ticker:
                        continue
                    precio_data = precios_todos_p.get(ticker)
                    if not precio_data:
                        continue
                    precio = precio_data['c']
                    pct_change = precio_data.get('pct_change')
                    if not precio or precio <= 0:
                        continue
                    precio_manual = obtener_precio_manual_monitor(bono.get('tipo_bono'), bono['nombre'])
                    precio_calculo = precio_manual if precio_manual is not None else precio
                    mat = bono.get('maturity')
                    if not mat:
                        continue
                    mat_date = mat.date() if hasattr(mat, 'date') else mat
                    dr = max((mat_date - fecha_hoy_p).days, 0)
                    if dr == 0:
                        continue
                    vf = bono.get('valor_final', 0) or 0
                    tna = (vf - precio_calculo) / precio_calculo / dr * 365 if precio_calculo > 0 else None
                    tea = (1 + (vf - precio_calculo) / precio_calculo) ** (365.0 / dr) - 1 if precio_calculo > 0 and dr > 0 else None
                    tem = (1 + tea) ** (1 / 12) - 1 if tea is not None else None
                    vm = dr / 365.0

                    # Pre-populate calculator default price
                    key_precio = f"precio_lecap_{bono['nombre']}"
                    if key_precio not in st.session_state:
                        st.session_state[key_precio] = float(precio)

                    filas_lecap.append({
                        'Activo': bono['nombre'],
                        'Vencimiento': mat_date.strftime('%d/%m/%Y'),
                        'Precio': precio_calculo,
                        '_precio_mercado': precio,
                        'TNA': round(tna * 100, 2) if tna is not None else None,
                        'TEM': round(tem * 100, 2) if tem is not None else None,
                        'Vida Media': round(vm, 2),
                        'Días Rem.': dr,
                        'Valor Final': round(vf, 4),
                        'Var. Diaria %': pct_change,
                    })

            if filas_lecap:
                st.markdown(TABLE_CSS, unsafe_allow_html=True)
                st.markdown('<div class="bond-wrap"><div class="bond-title">Lecaps &amp; Boncaps</div></div>', unsafe_allow_html=True)
                df_lec = pd.DataFrame(filas_lecap)
                df_lec_raw = df_lec.copy()

                # Cueva de rendimientos — TNA vs Días Remanente
                df_curva_lec = df_lec_raw[['Activo', 'Días Rem.', 'TNA']].dropna()
                df_curva_lec = df_curva_lec[df_curva_lec['TNA'] > 0]
                if len(df_curva_lec) >= 2:
                    x_lec = df_curva_lec['Días Rem.'].values.astype(float)
                    y_lec = df_curva_lec['TNA'].values.astype(float)
                    fig_lec = go.Figure()
                    fig_lec.add_trace(go.Scatter(
                        x=x_lec, y=y_lec,
                        mode='markers+text',
                        text=df_curva_lec['Activo'],
                        textposition='top center',
                        textfont=dict(size=10, color='#1a237e'),
                        marker=dict(size=9, color='#1a237e'),
                        name='Lecaps & Boncaps',
                        hovertemplate='<b>%{text}</b><br>Días Rem.: %{x:.0f}<br>TNA: %{y:.2f}%<extra></extra>',
                    ))
                    if len(df_curva_lec) >= 3:
                        import numpy as _np2
                        # Regresión logarítmica: y = a * ln(x) + b
                        log_x_lec = _np2.log(x_lec)
                        coeffs_lec = _np2.polyfit(log_x_lec, y_lec, 1)
                        x_line_lec = _np2.linspace(x_lec.min(), x_lec.max(), 200)
                        y_line_lec = coeffs_lec[0] * _np2.log(x_line_lec) + coeffs_lec[1]
                        fig_lec.add_trace(go.Scatter(
                            x=x_line_lec, y=y_line_lec,
                            mode='lines',
                            line=dict(color='#42a5f5', width=2, dash='dash'),
                            name='Tendencia',
                            hoverinfo='skip',
                            showlegend=False,
                        ))
                    fig_lec.update_layout(
                        title='Cueva de Rendimientos — Lecaps & Boncaps',
                        xaxis_title='Días al Vencimiento',
                        yaxis_title='TNA (%)',
                        xaxis=dict(showgrid=True, gridcolor='#cccccc', linecolor='#999999', linewidth=1, showline=True, tickfont=dict(color='#444444'), title_font=dict(color='#444444')),
                        yaxis=dict(showgrid=True, gridcolor='#cccccc', linecolor='#999999', linewidth=1, showline=True, tickfont=dict(color='#444444'), title_font=dict(color='#444444')),
                        plot_bgcolor='#f4f6fb',
                        paper_bgcolor='#f4f6fb',
                        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
                        height=420,
                        margin=dict(t=60, b=40, l=60, r=20),
                    )
                    st.plotly_chart(fig_lec, use_container_width=True)

                # Tabla
                df_lec = df_lec.sort_values('Días Rem.').reset_index(drop=True)
                df_lec_display = df_lec.drop(columns=['_precio_mercado'], errors='ignore').copy()
                tabla_id_lec = _tabla_manual_id('Lecaps & Boncaps')
                df_lec_display['_manual_control'] = df_lec['Activo'].apply(
                    lambda activo: (
                        f'<a href="?manual_table={tabla_id_lec}&manual_toggle={activo}&manual_market='
                        f'{float(df_lec.loc[df_lec["Activo"] == activo, "_precio_mercado"].iloc[0]):.6f}" '
                        'style="text-decoration:none;font-size:16px;color:#2e7d32">●</a>'
                        if _obtener_precio_manual_tabla(tabla_id_lec, activo) is not None else
                        f'<a href="?manual_table={tabla_id_lec}&manual_toggle={activo}&manual_market='
                        f'{float(df_lec.loc[df_lec["Activo"] == activo, "_precio_mercado"].iloc[0]):.6f}" '
                        'style="text-decoration:none;font-size:16px;color:#9aa5b1">○</a>'
                    )
                )
                df_lec_display['Precio'] = df_lec_display['Precio'].map('{:.2f}'.format)
                df_lec_display['TNA'] = df_lec_display['TNA'].apply(lambda v: f'{v:.2f}%' if v is not None else '-')
                df_lec_display['TEM'] = df_lec_display['TEM'].apply(lambda v: f'{v:.2f}%' if v is not None else '-')
                df_lec_display['Vida Media'] = df_lec_display['Vida Media'].map('{:.2f}'.format)
                df_lec_display['Valor Final'] = df_lec_display['Valor Final'].map('{:.4f}'.format)
                df_lec_display['Var. Diaria %'] = df_lec_display['Var. Diaria %'].apply(
                    lambda x: f'{x:+.2f}%' if x is not None and not pd.isna(x) else '-'
                )
                st.markdown("<div style='margin-top:2rem'></div>", unsafe_allow_html=True)
                st.markdown(render_tabla_html(df_lec_display), unsafe_allow_html=True)
            else:
                st.info("No hay precios disponibles en este momento.")

            # --- Tabla Bonos CER ---
            st.markdown("<div style='margin-top:2.5rem'></div>", unsafe_allow_html=True)

            # Obtener CER settlement desde BCRA (10° día hábil anterior a fecha_hoy_p)
            _cer_settl_tab, _ = obtener_cer_settlement(fecha_hoy_p)

            bonos_cer = [b for b in bonos if b.get('tipo_bono') == 'Bonos CER']
            filas_cer = []
            for bono in bonos_cer:
                ticker = bono.get('ticker', '').strip().upper()
                if not ticker:
                    continue
                precio_data = precios_todos_p.get(ticker)
                if not precio_data:
                    continue
                precio = precio_data['c']
                pct_change = precio_data.get('pct_change')
                if not precio or precio <= 0:
                    continue
                precio_manual = obtener_precio_manual_monitor(bono.get('tipo_bono'), bono['nombre'])
                precio_calculo = precio_manual if precio_manual is not None else precio
                mat = bono.get('maturity')
                if not mat:
                    continue
                mat_date = mat.date() if hasattr(mat, 'date') else mat
                dr = max((mat_date - fecha_hoy_p).days, 0)
                if dr == 0:
                    continue

                # Factor CER vivo = CER settlement / CER base
                _cb = bono.get('cer_base') or 0
                factor_cer_vivo = round(_cer_settl_tab / _cb, 4) if (_cer_settl_tab and _cb) else None

                # TIR Anual y TIR Mensual
                if factor_cer_vivo and precio_calculo > 0 and dr > 0:
                    tir_anual = (factor_cer_vivo * 100 / precio_calculo) ** (365.0 / dr) - 1
                    tir_mensual = (1 + tir_anual) ** (30 / 360) - 1
                else:
                    tir_anual = None
                    tir_mensual = None

                # Duración Modificada (bullet)
                vm = dr / 365.0
                dur_mod = vm / (1 + tir_anual) if (tir_anual is not None and vm > 0) else None

                # Pre-populate calculator default price
                key_precio_cer = f"precio_cer_{bono['nombre']}"
                if key_precio_cer not in st.session_state:
                    st.session_state[key_precio_cer] = float(precio)

                filas_cer.append({
                    'Activo': bono['nombre'],
                    'Vencimiento': mat_date.strftime('%d/%m/%Y'),
                    'Factor CER': factor_cer_vivo,
                    'Precio': precio_calculo,
                    '_precio_mercado': precio,
                    'TIR Anual': round(tir_anual * 100, 2) if tir_anual is not None else None,
                    'TIR Mensual': round(tir_mensual * 100, 2) if tir_mensual is not None else None,
                    'Dur. Modificada': round(dur_mod, 2) if dur_mod is not None else None,
                    'Var. Diaria %': pct_change,
                })

            if filas_cer:
                st.markdown(TABLE_CSS, unsafe_allow_html=True)
                st.markdown('<div class="bond-wrap"><div class="bond-title">Bonos CER</div></div>', unsafe_allow_html=True)
                df_cer = pd.DataFrame(filas_cer)
                df_cer_raw = df_cer.copy()

                # Cueva — TIR Anual vs Duración Modificada
                df_curva_cer = df_cer_raw[['Activo', 'Dur. Modificada', 'TIR Anual']].dropna()
                df_curva_cer = df_curva_cer[df_curva_cer['TIR Anual'].notna()]
                if len(df_curva_cer) >= 2:
                    x_cer = df_curva_cer['Dur. Modificada'].values.astype(float)
                    y_cer = df_curva_cer['TIR Anual'].values.astype(float)
                    fig_cer = go.Figure()
                    fig_cer.add_trace(go.Scatter(
                        x=x_cer, y=y_cer,
                        mode='markers+text',
                        text=df_curva_cer['Activo'],
                        textposition='top center',
                        textfont=dict(size=10, color='#1a237e'),
                        marker=dict(size=9, color='#1a237e'),
                        name='Bonos CER',
                        hovertemplate='<b>%{text}</b><br>Dur. Mod.: %{x:.2f}<br>TIR Anual: %{y:.2f}%<extra></extra>',
                    ))
                    if len(df_curva_cer) >= 3:
                        coeffs_cer = np.polyfit(np.log(x_cer), y_cer, 1)
                        x_line_cer = np.linspace(x_cer.min(), x_cer.max(), 200)
                        y_line_cer = coeffs_cer[0] * np.log(x_line_cer) + coeffs_cer[1]
                        fig_cer.add_trace(go.Scatter(
                            x=x_line_cer, y=y_line_cer,
                            mode='lines',
                            line=dict(color='#42a5f5', width=2, dash='dash'),
                            name='Tendencia',
                            hoverinfo='skip',
                            showlegend=False,
                        ))
                    fig_cer.update_layout(
                        title='Cueva de Rendimientos — Bonos CER',
                        xaxis_title='Duración Modificada (años)',
                        yaxis_title='TIR Anual (%)',
                        xaxis=dict(showgrid=True, gridcolor='#cccccc', linecolor='#999999', linewidth=1, showline=True, tickfont=dict(color='#444444'), title_font=dict(color='#444444')),
                        yaxis=dict(showgrid=True, gridcolor='#cccccc', linecolor='#999999', linewidth=1, showline=True, tickfont=dict(color='#444444'), title_font=dict(color='#444444')),
                        plot_bgcolor='#f4f6fb',
                        paper_bgcolor='#f4f6fb',
                        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='right', x=1),
                        height=420,
                        margin=dict(t=60, b=40, l=60, r=20),
                    )
                    st.plotly_chart(fig_cer, use_container_width=True)

                # Tabla
                df_cer = df_cer.sort_values('Dur. Modificada').reset_index(drop=True)
                df_cer_display = df_cer.drop(columns=['_precio_mercado'], errors='ignore').copy()
                tabla_id_cer = _tabla_manual_id('Bonos CER')
                df_cer_display['_manual_control'] = df_cer['Activo'].apply(
                    lambda activo: (
                        f'<a href="?manual_table={tabla_id_cer}&manual_toggle={activo}&manual_market='
                        f'{float(df_cer.loc[df_cer["Activo"] == activo, "_precio_mercado"].iloc[0]):.6f}" '
                        'style="text-decoration:none;font-size:16px;color:#2e7d32">●</a>'
                        if _obtener_precio_manual_tabla(tabla_id_cer, activo) is not None else
                        f'<a href="?manual_table={tabla_id_cer}&manual_toggle={activo}&manual_market='
                        f'{float(df_cer.loc[df_cer["Activo"] == activo, "_precio_mercado"].iloc[0]):.6f}" '
                        'style="text-decoration:none;font-size:16px;color:#9aa5b1">○</a>'
                    )
                )
                df_cer_display['Factor CER'] = df_cer_display['Factor CER'].apply(lambda v: f'{v:.4f}' if v is not None else '-')
                df_cer_display['Precio'] = df_cer_display['Precio'].map('{:.2f}'.format)
                df_cer_display['TIR Anual'] = df_cer_display['TIR Anual'].apply(lambda v: f'{v:.2f}%' if v is not None else '-')
                df_cer_display['TIR Mensual'] = df_cer_display['TIR Mensual'].apply(lambda v: f'{v:.2f}%' if v is not None else '-')
                df_cer_display['Dur. Modificada'] = df_cer_display['Dur. Modificada'].apply(lambda v: f'{v:.2f}' if v is not None else '-')
                df_cer_display['Var. Diaria %'] = df_cer_display['Var. Diaria %'].apply(
                    lambda x: f'{x:+.2f}%' if x is not None and not pd.isna(x) else '-'
                )
                st.markdown("<div style='margin-top:2rem'></div>", unsafe_allow_html=True)
                st.markdown(render_tabla_html(df_cer_display), unsafe_allow_html=True)
            else:
                st.markdown("<div style='margin-top:1rem'></div>", unsafe_allow_html=True)
                st.info("Bonos CER: no hay precios disponibles en este momento.")


except FileNotFoundError:
    st.error("❌ No se pudo cargar el archivo de datos")
    st.info("Asegúrese de que el archivo 'bonos_flujos.xlsx' esté en el directorio correcto")
except Exception as e:
    st.error(f"❌ Error al cargar los datos: {e}")
